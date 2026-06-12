from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from io import BytesIO

HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
HEADER_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
REQUIRED_FILL = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
OPTIONAL_FILL = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
SAMPLE_FILL = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
SAMPLE_FONT = Font(name="Calibri", italic=True, color="375623", size=10)
HELP_FONT = Font(name="Calibri", color="595959", size=9, italic=True)
NOTE_FILL = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)
FK_FILL = PatternFill(start_color="F4B942", end_color="F4B942", fill_type="solid")


def auto_width(ws, min_w=10, max_w=40):
    for col in ws.columns:
        mx = min_w
        for cell in col:
            if cell.value:
                mx = max(mx, min(len(str(cell.value)) + 2, max_w))
        ws.column_dimensions[get_column_letter(col[0].column)].width = mx


def write_sheet(ws, title, fields, samples, notes=None):
    ws.title = title

    headers = [f[0] for f in fields]
    types = [f[1] if len(f) > 1 else "" for f in fields]
    required = [f[2] if len(f) > 2 else True for f in fields]
    descs = [f[3] if len(f) > 3 else "" for f in fields]

    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN_BORDER

    for c, (t, r) in enumerate(zip(types, required), 1):
        is_fk = t.startswith("FK")
        label = f"[{'REQ' if r else 'OPC'}] {t}"
        cell = ws.cell(row=2, column=c, value=label)
        cell.font = Font(name="Calibri", size=8, bold=True, color="1F4E79")
        cell.fill = FK_FILL if is_fk else (REQUIRED_FILL if r else OPTIONAL_FILL)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = THIN_BORDER

    for c, d in enumerate(descs, 1):
        cell = ws.cell(row=3, column=c, value=d)
        cell.font = HELP_FONT
        cell.fill = NOTE_FILL
        cell.alignment = Alignment(wrap_text=True, vertical="top")
        cell.border = THIN_BORDER

    for r, sample in enumerate(samples, 4):
        for c, val in enumerate(sample, 1):
            cell = ws.cell(row=r, column=c, value=val)
            cell.font = SAMPLE_FONT
            cell.fill = SAMPLE_FILL
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = THIN_BORDER

    if notes:
        nr = len(samples) + 5
        ws.merge_cells(start_row=nr, start_column=1, end_row=nr, end_column=len(headers))
        cell = ws.cell(row=nr, column=1, value=notes)
        cell.font = HELP_FONT
        cell.fill = NOTE_FILL
        cell.alignment = Alignment(wrap_text=True)

    ws.freeze_panes = "A4"
    ws.sheet_properties.tabColor = "1F4E79"
    auto_width(ws)


def fk(desc):
    return f"{desc} (ID existente)"


BASE_DATE = "17/03/2025"


def _base_val(hour, minute, peak_low, peak_high, off_peak_low, off_peak_high):
    """interpolate between off-peak and peak values based on time proximity to peak."""
    minutes = hour * 60 + minute
    if 7 * 60 <= minutes <= 9 * 60:
        frac = (minutes - 7 * 60) / 120
        return int(peak_low + (peak_high - peak_low) * (1 - abs(frac - 0.5) * 2))
    if 18 * 60 <= minutes <= 20 * 60:
        frac = (minutes - 18 * 60) / 120
        return int(peak_low + (peak_high - peak_low) * (1 - abs(frac - 0.5) * 2))
    return (off_peak_low + off_peak_high) // 2


PC_MOV = {'PC-01': '12', 'PC-02': '21'}

def _periodizacion_rows():
    rows = []
    combos = [
        ("Estudio de Semaforos Av. Libertador", "PC-01", "PM-L", 6, 9),
        ("Estudio de Semaforos Av. Libertador", "PC-01", "PT-L", 18, 21),
        ("Estudio de Semaforos Av. Libertador", "PC-02", "PM-L", 6, 9),
        ("Estudio de Semaforos Av. Libertador", "PC-02", "PT-L", 18, 21),
        ("Auditoria Red Vial Centro", "PC-01", "PM-L", 6, 9),
        ("Auditoria Red Vial Centro", "PC-01", "PT-L", 18, 21),
        ("Auditoria Red Vial Centro", "PC-02", "PM-L", 6, 9),
        ("Auditoria Red Vial Centro", "PC-02", "PT-L", 18, 21),
    ]
    for proj, pc, period, h_start, h_end in combos:
        for hour in range(h_start, h_end):
            for minute in (0, 15, 30, 45):
                hora = f"{hour:02d}:{minute:02d}"
                vl = _base_val(hour, minute, 750, 1100, 300, 500)
                txc = _base_val(hour, minute, 80, 150, 30, 60)
                txb = _base_val(hour, minute, 25, 55, 10, 20)
                c2e = _base_val(hour, minute, 18, 40, 8, 15)
                c_mas2e = _base_val(hour, minute, 6, 18, 2, 5)
                peat = _base_val(hour, minute, 150, 350, 60, 120)
                cicl = _base_val(hour, minute, 25, 70, 10, 20)
                moto = _base_val(hour, minute, 35, 100, 15, 30)
                rows.append((BASE_DATE, hora, pc, period, proj, vl, txc, txb, c2e, c_mas2e, peat, cicl, moto))
    return rows


def generar_plantilla():
    wb = Workbook()

    ws_readme = wb.active
    ws_readme.title = "\U0001f4d6 README"

    readme_content = [
        ("PLANTILLA DE IMPORTACION \u2014 EIT App", ""),
        ("", ""),
        ("Cada hoja representa una tabla del sistema.", ""),
        ("Las columnas con fondo AZUL son REQUERIDAS (no pueden estar vacias).", ""),
        ("Las columnas con fondo GRIS son OPCIONALES.", ""),
        ("Las columnas con fondo NARANJO son claves foraneas (referencian otra tabla).", ""),
        ("", ""),
        ("INSTRUCCIONES:", ""),
        ("1. NO modifiques la fila 1 (cabeceras) ni la fila 2 (tipos).", ""),
        ("2. Reemplaza los datos de ejemplo (filas verde claro) con tus datos reales.", ""),
        ("3. Respeta el orden de las columnas.", ""),
        ("4. Respeta los formatos: fechas como DD/MM/AAAA, numeros sin separadores.", ""),
        ("5. Las columnas FK (naranjo) deben contener IDs existentes en el sistema.", ""),
        ("6. Puedes importar una sola hoja o varias a la vez.", ""),
        ("7. NO elimines hojas ni cabeceras \u2014 el importador las necesita.", ""),
        ("", ""),
        ("ORDEN RECOMENDADO DE CARGA:", ""),
        ("1. Mandante  \u2192  2. Contacto  \u2192  3. Proyecto  \u2192  4. Calle", ""),
        ("5. Nodo  \u2192  6. Arco  \u2192  7. Regulacion  \u2192  8. CoeficienteCruce", ""),
        ("9. Periodo  \u2192  10. PuntoControl  \u2192  11. Periodizacion", ""),
        ("12. ParametroArco  \u2192  13. FaseSemaforica  \u2192  14. ConfiguracionTransyt", ""),
        ("", ""),
        ("NOTA: Si importas todo de una vez, el sistema resolbera las FK automaticamente", ""),
        ("usando los identificadores unicos que proporciones.", ""),
        ("", ""),
        ("NOTA: Los datos de ejemplo en verde son consistentes entre todas las hojas", ""),
        ("y pueden importarse exitosamente como demostracion.", ""),
    ]

    for r, (text, _) in enumerate(readme_content, 1):
        cell = ws_readme.cell(row=r, column=1, value=text)
        if r == 1:
            cell.font = Font(name="Calibri", bold=True, size=16, color="1F4E79")
        elif text.startswith("INSTRUCCIONES") or text.startswith("ORDEN") or text.startswith("NOTA"):
            cell.font = Font(name="Calibri", bold=True, size=11, color="1F4E79")
        else:
            cell.font = Font(name="Calibri", size=11)

    ws_readme.column_dimensions["A"].width = 95
    ws_readme.sheet_properties.tabColor = "1F4E79"

    sheets_data = [
        ("Mandante", [
            ("name", "Texto (100)", True, "Nombre del mandante/cliente"),
            ("location", "Texto (100)", True, "Ubicacion o direccion"),
            ("details", "Texto", False, "Notas u observaciones"),
        ], [
            ("Municipalidad de Santiago", "Santiago Centro", "Comuna piloto"),
            ("Gobierno Regional Metropolitano", "Santiago", ""),
        ], "NOTA: El ID se genera automaticamente. El sistema asignara UUIDs."),

        ("Contacto", [
            ("name", "Texto (100)", True, "Nombre completo del contacto"),
            ("email", "Email", False, "Correo electronico"),
            ("phone", "Texto (20)", False, "Telefono / celular"),
            ("cargo", "Texto (100)", False, "Cargo o puesto"),
            ("position", "Texto (100)", False, "Posicion / departamento"),
            ("details", "Texto", False, "Notas"),
            ("mandante", "FK", True, fk("Mandante.name")),
        ], [
            ("Juan Perez", "jperez@msantiago.cl", "+56912345678", "Jefe de Transito", "Dpto. Vialidad", "", "Municipalidad de Santiago"),
            ("Maria Rojas", "", "", "Ingeniera", "Obras", "", "Gobierno Regional Metropolitano"),
        ], "NOTA: 'mandante' debe coincidir con el campo 'name' de la hoja Mandante."),

        ("Proyecto", [
            ("title", "Texto (100)", True, "Nombre del proyecto"),
            ("description", "Texto", False, "Descripcion detallada"),
            ("date_started", "Fecha (DD/MM/AAAA)", True, "Fecha de inicio"),
            ("mandante", "FK", True, fk("Mandante.name")),
        ], [
            ("Estudio de Semaforos Av. Libertador", "Actualizacion de planes de tiempo", "01/03/2025", "Municipalidad de Santiago"),
            ("Auditoria Red Vial Centro", "Levantamiento y analisis de flujos", "15/01/2025", "Gobierno Regional Metropolitano"),
        ], "NOTA: El usuario y fechas de completado se asignan en el sistema. 'mandante' debe existir en la hoja Mandante."),

        ("Calle", [
            ("nombre", "Texto (100)", True, "Nombre de la calle / avenida"),
            ("numero", "Entero", True, "Numero identificador unico por proyecto"),
            ("proyecto", "FK", True, fk("Proyecto.title")),
        ], [
            ("Av. Libertador", 1, "Estudio de Semaforos Av. Libertador"),
            ("Av. Providencia", 2, "Estudio de Semaforos Av. Libertador"),
            ("Av. Manuel Montt", 3, "Estudio de Semaforos Av. Libertador"),
            ("Av. Santa Maria", 1, "Auditoria Red Vial Centro"),
            ("Av. Los Leones", 2, "Auditoria Red Vial Centro"),
            ("Av. Tobalaba", 3, "Auditoria Red Vial Centro"),
        ], "NOTA: La combinacion 'numero + proyecto' debe ser unica."),

        ("Nodo", [
            ("numero", "Entero", True, "Numero del nodo (unico por proyecto)"),
            ("interseccion", "Texto (200)", False, "Descripcion de la interseccion"),
            ("calle_1", "FK (Calle.nombre)", False, fk("Calle.nombre")),
            ("calle_2", "FK (Calle.nombre)", False, fk("Calle.nombre")),
            ("numero_pc", "Entero (0-99)", False, "Numero del punto de control si aplica"),
            ("plano", "URL", False, "Link a imagen del plano"),
            ("imagen", "URL", False, "Link a foto del nodo"),
            ("proyecto", "FK", True, fk("Proyecto.title")),
        ], [
            (1, "Av. Libertador con Av. Providencia", "Av. Libertador", "Av. Providencia", 1, "", "", "Estudio de Semaforos Av. Libertador"),
            (2, "Av. Libertador con Av. Manuel Montt", "Av. Libertador", "Av. Manuel Montt", 2, "", "", "Estudio de Semaforos Av. Libertador"),
            (1, "Av. Santa Maria con Av. Los Leones", "Av. Santa Maria", "Av. Los Leones", 1, "", "", "Auditoria Red Vial Centro"),
            (2, "Av. Santa Maria con Av. Tobalaba", "Av. Santa Maria", "Av. Tobalaba", 2, "", "", "Auditoria Red Vial Centro"),
        ], "NOTA: 'calle_1' y 'calle_2' deben coincidir con 'nombre' en hoja Calle. 'numero_pc' asigna un PC al nodo."),

        ("Arco", [
            ("nodo_origen", "FK (Nodo.numero)", True, "Nodo de origen (debe existir en el proyecto)"),
            ("nodo_destino", "FK (Nodo.numero)", True, "Nodo de destino"),
            ("longitud", "Decimal", True, "Longitud en metros (ej: 120.5)"),
            ("proyecto", "FK", True, fk("Proyecto.title")),
        ], [
            (1, 2, 150.0, "Estudio de Semaforos Av. Libertador"),
            (2, 1, 150.0, "Estudio de Semaforos Av. Libertador"),
            (1, 2, 120.0, "Auditoria Red Vial Centro"),
            (2, 1, 120.0, "Auditoria Red Vial Centro"),
        ], "NOTA: 'nodo_origen' y 'nodo_destino' se resuelven por 'numero + proyecto'. La terna (origen, destino, proyecto) debe ser unica."),

        ("Regulacion", [
            ("codigo", "Texto (20)", True, "Codigo unico de regulacion (ej: SEM01, CEDA)"),
            ("descripcion", "Texto (100)", True, "Descripcion de la regulacion"),
        ], [
            ("SEM01", "Semaforo con etapa semaforica fija"),
            ("CEDA", "Ceda el paso"),
            ("PARE", "Señal de pare / stop"),
        ], "NOTA: 'codigo' debe ser unico en todo el sistema (no solo por proyecto)."),

        ("CoeficienteCruce", [
            ("nomenclatura", "Texto (10)", True, "Codigo corto (ej: VL, TXC, C2E)"),
            ("tipo_transporte", "Texto (50)", True, "Tipo de vehiculo (ej: Vehiculo Liviano, Taxi, Bus)"),
            ("coeficiente", "Decimal", True, "Factor de conversion a vehiculos equivalentes"),
            ("is_standard", "SI/NO", True, "¿Es estandar? 'SI'=global, 'NO'=solo para este proyecto"),
            ("proyecto", "FK", False, fk("Proyecto.title (solo si is_standard=NO)")),
        ], [
            ("VL", "Vehiculo Liviano", 1.0, "SI", ""),
            ("TXC", "Taxi Colectivo", 1.5, "SI", ""),
            ("BUS", "Bus", 2.5, "NO", "Estudio de Semaforos Av. Libertador"),
        ], "NOTA: Si 'is_standard' = SI, el coeficiente aplica a todos los proyectos. 'proyecto' solo se usa si NO es estandar."),

        ("Periodo", [
            ("codigo", "Texto (4)", True, "Codigo: PM-L / PN-L / PT-L / PE-L / PM-S / PN-S / PT-S / PE-S / PM-F / PN-F / PT-F / PE-F"),
            ("hora_inicio", "Hora HH:MM", False, "Hora de inicio (ej: 07:30)"),
            ("hora_fin", "Hora HH:MM", False, "Hora de termino (ej: 09:30)"),
            ("es_laboral", "SI/NO", True, "¿Es periodo laboral?"),
            ("proyecto", "FK", True, fk("Proyecto.title")),
        ], [
            ("PM-L", "06:00", "09:00", "SI", "Estudio de Semaforos Av. Libertador"),
            ("PT-L", "18:00", "21:00", "SI", "Estudio de Semaforos Av. Libertador"),
            ("PM-L", "06:00", "09:00", "SI", "Auditoria Red Vial Centro"),
            ("PT-L", "18:00", "21:00", "SI", "Auditoria Red Vial Centro"),
        ], "NOTA: Codigos: PM=Manana, PN=Mediodia, PT=Tarde, PE=Noche; L=Laboral, S=Sabado, F=Festivo. La combinacion (codigo, proyecto) debe ser unica."),

        ("PuntoControl", [
            ("nodo", "FK (Nodo.numero)", True, "Nodo asociado (numero dentro del proyecto)"),
            ("movimiento", "Texto (2)", True, "Codigo movimiento: 12,13,14,15,21,23,24,25,31,32,34,35,41,42,43,45,51,52,53,54"),
            ("viraje", "Texto (3)", False, "DIR=Directo, DER=Derecha, IZQ=Izquierda"),
            ("is_prioritario", "SI/NO", True, "¿Es punto de control prioritario?"),
            ("arco_entrada", "FK (Arco)", False, fk("Arco (nodo_origen > nodo_destino)")),
            ("arco_salida", "FK (Arco)", False, fk("Arco (nodo_origen > nodo_destino)")),
            ("regulacion", "FK (Regulacion.codigo)", False, "Codigo de regulacion (ej: SEM01)"),
            ("numero_pistas", "Decimal", False, "Numero de pistas (ej: 2.5)"),
            ("proyecto", "FK", True, fk("Proyecto.title")),
        ], [
            (1, "12", "DIR", "SI", "1>2", "2>1", "SEM01", 2.0, "Estudio de Semaforos Av. Libertador"),
            (2, "21", "DIR", "SI", "2>1", "1>2", "SEM01", 2.0, "Estudio de Semaforos Av. Libertador"),
            (1, "12", "DIR", "SI", "1>2", "2>1", "SEM01", 2.0, "Auditoria Red Vial Centro"),
            (2, "21", "DIR", "SI", "2>1", "1>2", "SEM01", 2.0, "Auditoria Red Vial Centro"),
        ], "NOTA: 'arco_entrada' y 'arco_salida' se indican como 'origen>destino' (ej: 1>2). 'movimiento' son 2 digitos: nodo_origen + nodo_destino."),

        ("Periodizacion", [
            ("fecha", "Fecha (DD/MM/AAAA)", True, "Fecha del conteo"),
            ("hora", "Hora HH:MM", True, "Hora del conteo"),
            ("pc", "FK (PuntoControl.nombre)", True, fk("Nombre del PC (ej: PC-01, PC-02)")),
            ("periodo", "FK (Periodo.codigo)", True, fk("Periodo.codigo (ej: PM-L)")),
            ("proyecto", "FK (Proyecto.title)", True, fk("Proyecto.title")),
            ("vl", "Entero", True, "Vehiculos Livianos"),
            ("txc", "Entero", True, "Taxi Colectivo"),
            ("txb", "Entero", True, "Taxi Bus"),
            ("c2e", "Entero", True, "Camion 2 ejes"),
            ("c_mas2e", "Entero", True, "Camion mas de 2 ejes"),
            ("peat", "Entero", True, "Peatones"),
            ("cicl", "Entero", True, "Ciclistas"),
            ("moto", "Entero", True, "Motocicletas"),
        ], _periodizacion_rows(), "NOTA: 'ftot' (flujo total) se calcula automaticamente. La combinacion (fecha, pc_mov, hora) debe ser unica."),

        ("ParametroArco", [
            ("punto_control", "FK", True, fk("PuntoControl (nombre: PC-XX o Nodo-XX)")),
            ("flujo_saturacion", "Decimal", True, "Flujo de saturacion en ADE/hr verde (ej: 1800)"),
            ("ponderador_demora", "Decimal", True, "Ponderador de demora (ej: 1.0)"),
            ("ponderador_detencion", "Decimal", True, "Ponderador de detencion (ej: 1.0)"),
            ("capacidad_cola", "Decimal", False, "Capacidad de cola en vehiculos"),
            ("tiene_tarjeta_38", "SI/NO", True, "¿Tiene tarjeta 38?"),
            ("proyecto", "FK", True, fk("Proyecto.title")),
        ], [
            ("PC-01", 1800.0, 1.0, 1.0, 10.0, "SI", "Estudio de Semaforos Av. Libertador"),
            ("PC-02", 1600.0, 1.2, 0.8, "", "NO", "Estudio de Semaforos Av. Libertador"),
            ("PC-01", 1750.0, 1.0, 1.0, 12.0, "SI", "Auditoria Red Vial Centro"),
            ("PC-02", 1550.0, 1.1, 0.9, "", "NO", "Auditoria Red Vial Centro"),
        ], "NOTA: Cada PuntoControl tiene UN ParametroArco (relacion 1 a 1)."),

        ("FaseSemaforica", [
            ("punto_control", "FK", True, fk("PuntoControl (nombre: PC-XX o Nodo-XX)")),
            ("fase_numero", "Entero", True, "Numero de fase (1, 2, 3...)"),
            ("verde_inicio", "Decimal", True, "Inicio del verde en segundos (ej: 0.0)"),
            ("verde_fin", "Decimal", True, "Fin del verde en segundos (ej: 25.0)"),
            ("proyecto", "FK", True, fk("Proyecto.title")),
        ], [
            ("PC-01", 1, 0.0, 25.0, "Estudio de Semaforos Av. Libertador"),
            ("PC-01", 2, 30.0, 55.0, "Estudio de Semaforos Av. Libertador"),
            ("PC-02", 1, 0.0, 20.0, "Estudio de Semaforos Av. Libertador"),
            ("PC-02", 2, 25.0, 45.0, "Estudio de Semaforos Av. Libertador"),
            ("PC-01", 1, 0.0, 30.0, "Auditoria Red Vial Centro"),
            ("PC-01", 2, 35.0, 60.0, "Auditoria Red Vial Centro"),
            ("PC-02", 1, 0.0, 22.0, "Auditoria Red Vial Centro"),
            ("PC-02", 2, 28.0, 50.0, "Auditoria Red Vial Centro"),
        ], "NOTA: La combinacion (punto_control, fase_numero) debe ser unica. Los tiempos son relativos al inicio del ciclo."),

        ("ConfiguracionTransyt", [
            ("proyecto", "FK", True, fk("Proyecto.title")),
            ("ciclo", "Entero", True, "Tiempo de ciclo en segundos (ej: 60)"),
            ("W", "Decimal", True, "Costo por hora de demora (ej: 10.0)"),
            ("K", "Decimal", True, "Costo por detencion (ej: 0.5)"),
            ("perdida_inicial", "Decimal", True, "Perdida inicial en segundos (ej: 2.0)"),
            ("ganancia_final", "Decimal", True, "Ganancia final en segundos (ej: 1.0)"),
        ], [
            ("Estudio de Semaforos Av. Libertador", 60, 10.0, 0.5, 2.0, 1.0),
            ("Auditoria Red Vial Centro", 70, 12.0, 0.4, 3.0, 1.5),
        ], "NOTA: Cada proyecto tiene UNA configuracion TRANSYT (relacion 1 a 1)."),
    ]

    for title, fields, samples, notes in sheets_data:
        ws = wb.create_sheet()
        write_sheet(ws, title, fields, samples, notes)

    return wb


def generar_plantilla_bytes():
    wb = generar_plantilla()
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf
