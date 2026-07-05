"""
Script de migración: origen.xlsx → destino.xlsx (plantilla EIT App)

Fuentes (3 primeras hojas de origen):
  1. Esquema de Intersecciones  → Calle, Nodo
  2. Periodización              → Periodo, Periodizacion
  3. Resumen Flujos             → Arco, PuntoControl, ParametroArco

Hoja nueva agregada al destino: 'Proyecto'
  → Se usa como fuente para Mandante, Contacto y Proyecto.
  → Completar esa hoja antes de importar al sistema.

Uso:
  python migrar_origen_a_destino.py [--origen origen.xlsx] [--destino destino.xlsx]
                                    [--salida output.xlsx] [--FP dd/mm/aaaa]
"""

import argparse
from copy import copy
from io import BytesIO

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill

# ── constantes ─────────────────────────────────────────────────────────────────

DEFAULT_FECHA = "01/01/2024"
PLACEHOLDER_PRY = "PROYECTO_PLACEHOLDER"
DATA_START_ROW = 3  # fila 1=headers, fila 2=tipos, datos desde fila 3
DATA_FILL = PatternFill("solid", start_color="D4E8D0", end_color="D4E8D0")


# ── helpers ────────────────────────────────────────────────────────────────────


def es_valido(v):
    return v is not None and str(v).strip() not in ("", "nan", "-", "#REF!")


def limpiar(s):
    return str(s).strip() if es_valido(s) else None


def validar_codigo_arco(val, col_name):
    if not es_valido(val):
        return False, f"{col_name} en blanco"
    s = str(val).strip()
    if not s.isdigit():
        return False, f"{col_name} '{s}' no es un numero"
    if len(s) < 3:
        return False, f"{col_name} '{s}' tiene menos de 3 caracteres"
    if not s.endswith("1"):
        return False, f"{col_name} '{s}' no termina en 1"
    return True, ""


def decode_arco_nodos(arco_str, known_nodes):
    """Decodifica un código de arco usando la fórmula Excel:

        =LEFT(F2;QUOTIENT(LEN(F2)-1;2)) & ">" & MID(F2;QUOTIENT(LEN(F2)-1;2)+1;LEN(F2)-QUOTIENT(LEN(F2)-1;2)-1)

    El split es puramente posicional:
      - 3 dígitos (A B 1) → A > B
      - 4 dígitos (A B C 1) → A > BC
    El último dígito (siempre 1) se descarta.

    Retorna (nodo_1, nodo_2) o (None, None).
    known_nodes se usa solo para decidir si un nodo existe o hay que crearlo (en build_arcos).
    """
    if not es_valido(arco_str):
        return None, None
    s = str(arco_str).strip()
    if len(s) < 3:
        return None, None
    split = (len(s) - 1) // 2  # QUOTIENT(LEN-1, 2)
    return str(int(s[:split])), str(int(s[split:-1]))


# ── lectura de origen ──────────────────────────────────────────────────────────


def leer_esquema(path):
    ei = pd.read_excel(path, sheet_name="Esquema de Intersecciones", header=None)
    rows = ei.iloc[1:, 7:13].copy()
    rows.columns = ["N", "Nodo", "Calle1", "Calle2", "Interseccion", "Movimientos"]
    rows = rows.dropna(subset=["Nodo"])
    rows["Nodo"] = rows["Nodo"].astype(int)
    rows["Calle1"] = rows["Calle1"].apply(limpiar)
    rows["Calle2"] = rows["Calle2"].apply(limpiar)
    rows["Interseccion"] = rows["Interseccion"].apply(limpiar)
    calles = set(c for c in list(rows["Calle1"]) + list(rows["Calle2"]) if c)
    return rows, calles


def leer_resumen_flujos(path):
    rf = pd.read_excel(path, sheet_name="Resumen Flujos", header=None)
    data = rf.iloc[3:].copy()

    # tabla principal (cols 0-9)
    mov = data[[0, 1, 2, 3, 4, 5, 6, 7, 8, 9]].copy()
    mov.columns = [
        "PC",
        "LUGAR",
        "MOV",
        "VIR",
        "PS",
        "arco_id",
        "Dist",
        "llega_a",
        "Flujo",
        "Codigo",
    ]
    mov = mov.dropna(subset=["PC"])
    # #REF! y '-' → None
    for col in ["arco_id", "Dist", "llega_a"]:
        mov[col] = mov[col].apply(lambda v: None if not es_valido(v) else str(v).strip())
    mov = mov[mov["arco_id"].notna() & mov["llega_a"].notna()]

    # arcos únicos (nodo_origen y nodo_destino en nomenclatura TRANSYT)
    arcos = mov[["arco_id", "Dist", "llega_a"]].copy()
    arcos.columns = ["nodo_origen", "longitud", "nodo_destino"]
    arcos["longitud"] = pd.to_numeric(arcos["longitud"], errors="coerce")  # #REF! → NaN
    arcos = arcos.drop_duplicates(subset=["nodo_origen", "nodo_destino"])

    # tabla parámetros embebida (cols 15-22)
    param = data[[15, 16, 17, 18, 19, 20, 21, 22]].copy()
    param.columns = ["PC", "MOV", "VIR", "Arco", "Tipo", "Capacidad", "Nr_Pistas", "Vel_ini"]
    param = param.dropna(subset=["PC"])

    return mov, arcos, param


def leer_periodizacion(path):
    peri = pd.read_excel(path, sheet_name="Periodización", header=None)
    data = peri.iloc[2:].copy()
    data.columns = [
        "PC",
        "INTERSECCION",
        "HORA",
        "MOV",
        "PER",
        "VL",
        "TXC",
        "TXB",
        "C2E",
        "C_MAS2E",
        "PEAT",
        "CICL",
        "MOTO",
        "FTOT",
        "PERIODO",
        "c15",
        "c16",
        "c17",
        "HORA2",
        "FLUJO_15MIN",
        "FLUJO_HMOVIL",
        "PERIODO2",
    ]
    data = data.dropna(subset=["PC"])
    for col in ["VL", "TXC", "TXB", "C2E", "C_MAS2E", "PEAT", "CICL", "MOTO"]:
        data[col] = pd.to_numeric(data[col], errors="coerce").fillna(0).astype(int)
    return data


# ── construcción de tablas destino ─────────────────────────────────────────────


def build_calles(calles_set):
    rows = [
        {"nombre": n, "numero": i, "proyecto": PLACEHOLDER_PRY}
        for i, n in enumerate(sorted(calles_set), 1)
    ]
    return pd.DataFrame(rows)


def build_nodos(nodos_df):
    # Mapeo PC → numero de nodo (del Esquema de Intersecciones, col N° = numero_pc)
    pc_map = {}
    for _, r in nodos_df.iterrows():
        n_pc = int(r["N"]) if es_valido(r["N"]) else None
        if n_pc:
            pc_map[f"PC-{n_pc:02d}"] = r["Nodo"]

    rows = []
    for _, r in nodos_df.iterrows():
        n_pc = int(r["N"]) if es_valido(r["N"]) else None
        rows.append(
            {
                "numero": r["Nodo"],
                "interseccion": r["Interseccion"],
                "calle_1": r["Calle1"],
                "calle_2": r["Calle2"],
                "numero_pc": n_pc,
                "plano": None,
                "imagen": None,
                "proyecto": PLACEHOLDER_PRY,
            }
        )
    return pd.DataFrame(rows), pc_map


def build_arcos(arcos_df, known_nodes):
    new_nodes = set()
    errors = []
    rows = []
    for idx, r in arcos_df.iterrows():
        origen_raw = r["nodo_origen"]
        destino_raw = r["nodo_destino"]

        ok, motivo = validar_codigo_arco(origen_raw, "arco_id")
        if not ok:
            errors.append(f"Resumen Flujos fila {idx + 1} no valida por {motivo}")
            continue

        ok, motivo = validar_codigo_arco(destino_raw, "llega_a")
        if not ok:
            errors.append(f"Resumen Flujos fila {idx + 1} no valida por {motivo}")
            continue

        dec_o = decode_arco_nodos(origen_raw, known_nodes)
        if dec_o and dec_o[0]:
            nodo_origen = dec_o[0]
            if dec_o[1] not in known_nodes:
                new_nodes.add(dec_o[1])
        else:
            nodo_origen = origen_raw

        dec_d = decode_arco_nodos(destino_raw, known_nodes)
        if dec_d and dec_d[0]:
            nodo_destino = dec_d[0]
            if dec_d[1] not in known_nodes:
                new_nodes.add(dec_d[1])
        else:
            nodo_destino = destino_raw

        rows.append(
            {
                "nodo_origen": nodo_origen,
                "nodo_destino": nodo_destino,
                "longitud": r["longitud"] if pd.notna(r["longitud"]) else 1,
                "proyecto": PLACEHOLDER_PRY,
            }
        )

    # deduplicar por par (origen, destino)
    df = pd.DataFrame(rows)
    if not df.empty:
        dedup = []
        for (_ori, _des), grp in df.groupby(["nodo_origen", "nodo_destino"], sort=False):
            row = grp.iloc[0].to_dict()
            if len(grp) > 1 and grp["longitud"].nunique() > 1:
                row["longitud"] = 1
            dedup.append(row)
        df = pd.DataFrame(dedup)
    return df, new_nodes, errors


def ensure_arcs_for_pc(pc_out, arcos_out, nodos_out, known_nodes):
    existing = set()
    if not arcos_out.empty:
        for _, r in arcos_out.iterrows():
            existing.add((str(r["nodo_origen"]), str(r["nodo_destino"])))

    missing = set()
    new_nodes_needed = set()
    for _, r in pc_out.iterrows():
        for col in ("arco_entrada", "arco_salida"):
            val = r.get(col)
            if isinstance(val, str) and ">" in val:
                parts = val.split(">", 1)
                key = (parts[0], parts[1])
                if key not in existing:
                    missing.add(key)
                    existing.add(key)
                    new_nodes_needed.add(parts[0])
                    new_nodes_needed.add(parts[1])

    # Ensure "1>1" sentinel arco exists (for default when PC arco is missing)
    if ("1", "1") not in existing:
        missing.add(("1", "1"))
        existing.add(("1", "1"))
        new_nodes_needed.add("1")

    if missing:
        for ori, des in missing:
            arcos_out = pd.concat(
                [
                    arcos_out,
                    pd.DataFrame(
                        [
                            {
                                "nodo_origen": ori,
                                "nodo_destino": des,
                                "longitud": 1,
                                "proyecto": PLACEHOLDER_PRY,
                            }
                        ]
                    ),
                ],
                ignore_index=True,
            )
        for n_str in new_nodes_needed:
            if n_str not in known_nodes and n_str.isdigit():
                nodos_out = pd.concat(
                    [
                        nodos_out,
                        pd.DataFrame(
                            [
                                {
                                    "numero": int(n_str),
                                    "interseccion": None,
                                    "calle_1": None,
                                    "calle_2": None,
                                    "numero_pc": None,
                                    "plano": None,
                                    "imagen": None,
                                    "proyecto": PLACEHOLDER_PRY,
                                }
                            ]
                        ),
                    ],
                    ignore_index=True,
                )
                known_nodes.add(n_str)

    return arcos_out, nodos_out, len(missing)


def renombrar_calles(calles_out, nodos_out):
    calle_map = {}
    for i, (_, r) in enumerate(calles_out.iterrows(), 1):
        calle_map[r["nombre"]] = f"M{i}"
    calles_out["nombre"] = [f"M{i}" for i in range(1, len(calles_out) + 1)]

    calle_1_new = []
    calle_2_new = []
    next_m = len(calles_out) + 1
    for _, r in nodos_out.iterrows():
        c1 = r["calle_1"]
        c2 = r["calle_2"]
        if pd.notna(c1) and c1 in calle_map:
            calle_1_new.append(calle_map[c1])
        elif pd.isna(c1):
            name = f"M{next_m}"
            calle_1_new.append(name)
            calles_out = pd.concat(
                [
                    calles_out,
                    pd.DataFrame(
                        [
                            {
                                "nombre": name,
                                "numero": next_m,
                                "proyecto": PLACEHOLDER_PRY,
                            }
                        ]
                    ),
                ],
                ignore_index=True,
            )
            next_m += 1
        else:
            calle_1_new.append(c1)
        if pd.notna(c2) and c2 in calle_map:
            calle_2_new.append(calle_map[c2])
        elif pd.isna(c2) and pd.isna(r["calle_1"]):
            name = f"M{next_m}"
            calle_2_new.append(name)
            calles_out = pd.concat(
                [
                    calles_out,
                    pd.DataFrame(
                        [
                            {
                                "nombre": name,
                                "numero": next_m,
                                "proyecto": PLACEHOLDER_PRY,
                            }
                        ]
                    ),
                ],
                ignore_index=True,
            )
            next_m += 1
        elif pd.isna(c2):
            calle_2_new.append(None)
        else:
            calle_2_new.append(c2)

    nodos_out = nodos_out.copy()
    nodos_out["calle_1"] = calle_1_new
    nodos_out["calle_2"] = calle_2_new
    return calles_out, nodos_out


def normalizar_pc(val):
    if not val or not isinstance(val, str):
        return val
    import re

    m = re.search(r"PC\s*-?\s*(\d+)", val, re.IGNORECASE)
    if m:
        return f"PC-{int(m.group(1)):02d}"
    return val


def build_puntos_control(param_df, mov_df, pc_map, known_nodes):
    """
    arco_entrada = arco_id > nodo_del_PC  (col F = arco que entra al PC)
    arco_salida  = nodo_del_PC > llega_a  (col H = nodo al que llega el movimiento)
    """
    # índice mov_df por (PC, MOV) para obtener llega_a
    mov_idx = {}
    for _, r in mov_df.iterrows():
        key = (str(r["PC"]).strip(), str(r["MOV"]).strip())
        mov_idx[key] = limpiar(r["llega_a"])

    rows = []
    seen = set()
    for _, r in param_df.iterrows():
        pc_raw = limpiar(r["PC"])
        mov = limpiar(r["MOV"])
        if not pc_raw or not mov:
            continue
        pc = normalizar_pc(pc_raw)
        key = (pc_raw, mov)
        if key in seen:
            continue
        seen.add(key)

        vir = limpiar(r["VIR"])
        tipo = limpiar(r["Tipo"]) or ""
        is_pri = "SI" if "riorit" in tipo else "NO"
        arco_f = limpiar(r["Arco"])  # col F = arco de entrada (nodo_origen del arco)
        n_pc = pc_map.get(pc)  # numero de nodo del PC
        llega = mov_idx.get(key)  # col H = llega_a

        dec_a = decode_arco_nodos(arco_f, known_nodes)
        dec_l = decode_arco_nodos(llega, known_nodes)
        arco_entrada = f"{dec_a[0]}>{dec_a[1]}" if (dec_a[0] and dec_a[1]) else "1>1"
        arco_salida = f"{dec_l[0]}>{dec_l[1]}" if (dec_l[0] and dec_l[1]) else "1>1"

        rows.append(
            {
                "nodo": n_pc,
                "movimiento": mov,
                "viraje": vir,
                "is_prioritario": is_pri,
                "arco_entrada": arco_entrada,
                "arco_salida": arco_salida,
                "regulacion": None,
                "numero_pistas": r["Nr_Pistas"] if es_valido(r["Nr_Pistas"]) else None,
                "proyecto": PLACEHOLDER_PRY,
            }
        )
    return pd.DataFrame(rows)


def build_parametros_arco(param_df):
    rows = []
    seen = set()
    for _, r in param_df.iterrows():
        pc_raw = limpiar(r["PC"])
        if not pc_raw or pc_raw in seen:
            continue
        seen.add(pc_raw)
        pc = normalizar_pc(pc_raw)
        cap = r["Capacidad"] if es_valido(r["Capacidad"]) else None
        rows.append(
            {
                "punto_control": pc,
                "flujo_saturacion": cap,
                "ponderador_demora": 1.0,
                "ponderador_detencion": 1.0,
                "capacidad_cola": None,
                "tiene_tarjeta_38": "NO",
                "proyecto": PLACEHOLDER_PRY,
            }
        )
    return pd.DataFrame(rows)


def build_periodos(peri_df):
    horas_map = {
        "PM-L": ("06:00", "09:00", "SI"),
        "PT-L": ("17:00", "21:00", "SI"),
        "PN-L": ("12:00", "14:00", "SI"),
        "PE-L": ("21:00", "23:00", "SI"),
        "PM-S": ("06:00", "09:00", "NO"),
        "PT-S": ("17:00", "21:00", "NO"),
        "PM-F": ("06:00", "09:00", "NO"),
        "PT-F": ("17:00", "21:00", "NO"),
    }
    rows = []
    for cod in peri_df["PER"].dropna().unique():
        h_ini, h_fin, lab = horas_map.get(cod, ("00:00", "00:00", "SI"))
        rows.append(
            {
                "codigo": cod,
                "hora_inicio": h_ini,
                "hora_fin": h_fin,
                "es_laboral": lab,
                "proyecto": PLACEHOLDER_PRY,
            }
        )
    return pd.DataFrame(rows)


def build_periodizacion(peri_df, fecha):
    df = peri_df.copy()
    pc_ser = df["PC"].apply(lambda v: normalizar_pc(limpiar(v)))
    mov_ser = df["MOV"].apply(lambda v: str(v).strip() if es_valido(v) else "")
    df["pc_mov"] = pc_ser + "-" + mov_ser
    df["periodo"] = df["PER"].apply(limpiar)
    df["hora"] = df["HORA"].apply(
        lambda v: (
            v.strftime("%H:%M")
            if hasattr(v, "strftime")
            else (str(v).strip() if es_valido(v) else None)
        )
    )
    df["fecha"] = fecha
    df["proyecto"] = PLACEHOLDER_PRY
    count_cols = ["VL", "TXC", "TXB", "C2E", "C_MAS2E", "PEAT", "CICL", "MOTO"]
    group_cols = ["fecha", "hora", "pc_mov", "periodo", "proyecto"]
    grouped = df.groupby(group_cols, dropna=False)[count_cols].sum().reset_index()
    grouped.columns = [c.lower() for c in grouped.columns]
    return grouped


# ── hoja ProyectoMigrado (nueva) ─────────────────────────────────────────────

PROYECTO_HEADERS = [
    # (columna, descripcion, requerido, tipo)
    ("mandante_name", "Nombre del mandante/cliente", True, "Texto (100)"),
    ("mandante_location", "Ubicación o dirección del mandante", True, "Texto (100)"),
    ("mandante_details", "Notas u observaciones del mandante", False, "Texto"),
    ("contacto_name", "Nombre completo del contacto", True, "Texto (100)"),
    ("contacto_email", "Correo electrónico del contacto", False, "Email"),
    ("contacto_phone", "Teléfono / celular del contacto", False, "Texto (20)"),
    ("contacto_cargo", "Cargo o puesto del contacto", False, "Texto (100)"),
    ("contacto_position", "Posición / departamento del contacto", False, "Texto (100)"),
    (
        "proyecto_title",
        "Nombre del proyecto  ← reemplaza PROYECTO_PLACEHOLDER",
        True,
        "Texto (100)",
    ),
    ("proyecto_description", "Descripción detallada del proyecto", False, "Texto"),
    (
        "proyecto_date_started",
        "Fecha de inicio del proyecto (DD/MM/AAAA)",
        True,
        "Fecha DD/MM/AAAA",
    ),
]

COLOR_REQ = "BDD7EE"
COLOR_OPC = "D9D9D9"
COLOR_NOTA = "FFF2CC"


def crear_hoja_proyecto_migrado(wb, proyecto_data):
    """Crea hoja ProyectoMigrado (data desde fila 4, sin título PLANTILLA)
    y popula las hojas Proyecto y Mandante del template estándar."""
    if "ProyectoMigrado" in wb.sheetnames:
        del wb["ProyectoMigrado"]
    ws = wb.create_sheet("ProyectoMigrado", 1)

    for col_idx, (campo, desc, req, tipo) in enumerate(PROYECTO_HEADERS, 1):
        c1 = ws.cell(row=1, column=col_idx, value=campo)
        c2 = ws.cell(row=2, column=col_idx, value=f"[{'REQ' if req else 'OPC'}] {tipo}")
        c3 = ws.cell(row=3, column=col_idx, value=desc)
        color = COLOR_REQ if req else COLOR_OPC
        fill = PatternFill("solid", start_color=color, end_color=color)
        for c in (c1, c2, c3):
            c.fill = fill
            c.font = Font(bold=(c.row == 1))
        ws.column_dimensions[c1.column_letter].width = max(len(campo), len(desc)) * 0.9 + 4

    for col_idx, (campo, _desc, _req, _tipo) in enumerate(PROYECTO_HEADERS, 1):
        val = proyecto_data.get(campo, "")
        c = ws.cell(row=4, column=col_idx, value=val if val else "")
        c.fill = copy(DATA_FILL)
        c.alignment = Alignment(horizontal="center", vertical="center")

    if "Proyecto" in wb.sheetnames:
        escribir_hoja(
            wb["Proyecto"],
            pd.DataFrame(
                [
                    {
                        "title": proyecto_data.get("proyecto_title", ""),
                        "description": proyecto_data.get("proyecto_description", ""),
                        "date_started": proyecto_data.get("proyecto_date_started", ""),
                        "mandante": proyecto_data.get("mandante_name", ""),
                    }
                ]
            ),
        )

    if "Mandante" in wb.sheetnames:
        escribir_hoja(
            wb["Mandante"],
            pd.DataFrame(
                [
                    {
                        "name": proyecto_data.get("mandante_name", ""),
                        "location": proyecto_data.get("mandante_location", ""),
                        "details": proyecto_data.get("mandante_details", ""),
                    }
                ]
            ),
        )


# ── escritura en plantilla ─────────────────────────────────────────────────────


def escribir_hoja(ws, df):
    max_col = ws.max_column

    nota_info = None
    for m in list(ws.merged_cells.ranges):
        if m.min_row >= DATA_START_ROW:
            val = ws.cell(row=m.min_row, column=m.min_col).value
            if val and isinstance(val, str) and val.strip().startswith("NOTA:"):
                nota_info = (m.min_row, m.min_col, m.max_row, m.max_col, val)

    for m in list(ws.merged_cells.ranges):
        if m.min_row >= DATA_START_ROW:
            ws.unmerge_cells(str(m))

    if nota_info:
        nr, _, _, _, _ = nota_info
        ws.cell(row=nr, column=1).value = None

    last_row = DATA_START_ROW - 1
    for row in range(DATA_START_ROW, ws.max_row + 1):
        if any(ws.cell(row=row, column=c).value is not None for c in range(1, max_col + 1)):
            last_row = row

    start_row = last_row + 1
    if start_row < 4:
        start_row = 4

    center = Alignment(horizontal="center", vertical="center")
    for i, (_, fila) in enumerate(df.iterrows()):
        for j, col in enumerate(df.columns):
            val = fila[col]
            if not isinstance(val, str) and pd.isna(val):
                val = None
            c = ws.cell(row=start_row + i, column=j + 1, value=val)
            c.fill = copy(DATA_FILL)
            c.alignment = center

    if nota_info:
        _, _, _, _, nt = nota_info
        nota_row = start_row + len(df) + 1
        ws.merge_cells(start_row=nota_row, start_column=1, end_row=nota_row, end_column=max_col)
        c = ws.cell(row=nota_row, column=1, value=nt)
        c.font = Font(italic=True)
        c.fill = PatternFill(start_color=COLOR_NOTA, end_color=COLOR_NOTA, fill_type="solid")


# ── main ───────────────────────────────────────────────────────────────────────


def run(origen, destino, salida, fecha):
    print("Leyendo origen...")
    nodos_df, calles_set = leer_esquema(origen)
    mov_df, arcos_df, param_df = leer_resumen_flujos(origen)
    peri_df = leer_periodizacion(origen)

    print("Construyendo tablas...")
    calles_out = build_calles(calles_set)
    nodos_out, pc_map = build_nodos(nodos_df)
    known_nodes = set(str(n) for n in nodos_out["numero"])
    arcos_out, new_nodes, arco_errors = build_arcos(arcos_df, known_nodes)
    for node_str in new_nodes:
        nodos_out = pd.concat(
            [
                nodos_out,
                pd.DataFrame(
                    [
                        {
                            "numero": int(node_str),
                            "interseccion": None,
                            "calle_1": None,
                            "calle_2": None,
                            "numero_pc": None,
                            "plano": None,
                            "imagen": None,
                            "proyecto": PLACEHOLDER_PRY,
                        }
                    ]
                ),
            ],
            ignore_index=True,
        )
    pc_out = build_puntos_control(param_df, mov_df, pc_map, known_nodes)
    known_nodes = set(str(n) for n in nodos_out["numero"])
    arcos_out, nodos_out, pc_arcos_count = ensure_arcs_for_pc(
        pc_out, arcos_out, nodos_out, known_nodes
    )
    calles_out, nodos_out = renombrar_calles(calles_out, nodos_out)
    param_out = build_parametros_arco(param_df)
    periodos_out = build_periodos(peri_df)
    periodiz_out = build_periodizacion(peri_df, fecha)

    proyecto_data = {
        "mandante_name": "Mandante Migrado",
        "mandante_location": "",
        "mandante_details": "",
        "contacto_name": "",
        "contacto_email": "",
        "contacto_phone": "",
        "contacto_cargo": "",
        "contacto_position": "",
        "proyecto_title": "Proyecto Migrado",
        "proyecto_description": "",
        "proyecto_date_started": fecha,
    }

    print("Escribiendo en plantilla destino...")
    wb = load_workbook(destino)

    crear_hoja_proyecto_migrado(wb, proyecto_data)

    escribir_hoja(wb["Calle"], calles_out)
    escribir_hoja(wb["Nodo"], nodos_out)
    escribir_hoja(wb["Arco"], arcos_out)
    escribir_hoja(wb["PuntoControl"], pc_out)
    escribir_hoja(wb["ParametroArco"], param_out)
    escribir_hoja(wb["Periodo"], periodos_out)
    escribir_hoja(wb["Periodizacion"], periodiz_out)

    wb.save(salida)
    print(f"\n✅  Archivo generado: {salida}")

    print("\n── RESUMEN ────────────────────────────────────────────────")
    print(f"  Hoja ProyectoMigrado: {len(proyecto_data)} campos")
    print(f"  Calle           : {len(calles_out)} filas")
    print(f"  Nodo            : {len(nodos_out)} filas  ({len(new_nodes)} agregados desde arcos)")
    print(
        f"  Arco            : {len(arcos_out)} filas  "
        f"({(arcos_out['longitud'] == 1).sum()} con longitud default 1)"
        f"{f'  ({pc_arcos_count} desde PC)' if pc_arcos_count else ''}"
    )
    print(f"  PuntoControl    : {len(pc_out)} filas")
    print(f"  ParametroArco   : {len(param_out)} filas")
    print(f"  Periodo         : {len(periodos_out)} filas")
    print(f"  Periodizacion   : {len(periodiz_out)} filas  (fecha usada: {fecha})")

    if arco_errors:
        print(f"\n── ERRORES ({len(arco_errors)}) ─────────────────────────────────")
        for err in arco_errors:
            print(f"  ⚠  {err}")


def reemplazar_placeholder(wb, proyecto_title):
    """
    Reemplaza PROYECTO_PLACEHOLDER por proyecto_title en todas las celdas
    de todas las hojas del workbook.
    """
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                if cell.value and isinstance(cell.value, str) and PLACEHOLDER_PRY in cell.value:
                    cell.value = cell.value.replace(PLACEHOLDER_PRY, proyecto_title)


# ── entry point para uso desde web ──────────────────────────────────────────


def run_from_bytes(origen_bio, destino_bio, fecha, proyecto_data=None):
    """
    Igual que run() pero recibe BytesIO de origen y destino, y retorna un
    tuple (BytesIO, dict) con el resultado y estadísticas.
    """
    raw = origen_bio.read()
    nodos_df, calles_set = leer_esquema(BytesIO(raw))
    mov_df, arcos_df, param_df = leer_resumen_flujos(BytesIO(raw))
    peri_df = leer_periodizacion(BytesIO(raw))

    calles_out = build_calles(calles_set)
    nodos_out, pc_map = build_nodos(nodos_df)
    known_nodes = set(str(n) for n in nodos_out["numero"])
    arcos_out, new_nodes, arco_errors = build_arcos(arcos_df, known_nodes)
    for node_str in new_nodes:
        nodos_out = pd.concat(
            [
                nodos_out,
                pd.DataFrame(
                    [
                        {
                            "numero": int(node_str),
                            "interseccion": None,
                            "calle_1": None,
                            "calle_2": None,
                            "numero_pc": None,
                            "plano": None,
                            "imagen": None,
                            "proyecto": PLACEHOLDER_PRY,
                        }
                    ]
                ),
            ],
            ignore_index=True,
        )
    pc_out = build_puntos_control(param_df, mov_df, pc_map, known_nodes)
    known_nodes = set(str(n) for n in nodos_out["numero"])
    arcos_out, nodos_out, pc_arcos_count = ensure_arcs_for_pc(
        pc_out, arcos_out, nodos_out, known_nodes
    )
    calles_out, nodos_out = renombrar_calles(calles_out, nodos_out)
    param_out = build_parametros_arco(param_df)
    periodos_out = build_periodos(peri_df)
    periodiz_out = build_periodizacion(peri_df, fecha)

    stats = {
        "Calle": len(calles_out),
        "Nodo": len(nodos_out),
        "Arco": len(arcos_out),
        "PuntoControl": len(pc_out),
        "ParametroArco": len(param_out),
        "Periodo": len(periodos_out),
        "Periodizacion": len(periodiz_out),
        "ArcoErrores": arco_errors,
        "ArcoDesdePC": pc_arcos_count,
    }

    wb = load_workbook(destino_bio)
    proyecto_title = PLACEHOLDER_PRY
    if proyecto_data:
        proyecto_title = proyecto_data.get("proyecto_title", PLACEHOLDER_PRY)
        crear_hoja_proyecto_migrado(wb, proyecto_data)

    escribir_hoja(wb["Calle"], calles_out)
    escribir_hoja(wb["Nodo"], nodos_out)
    escribir_hoja(wb["Arco"], arcos_out)
    escribir_hoja(wb["PuntoControl"], pc_out)
    escribir_hoja(wb["ParametroArco"], param_out)
    escribir_hoja(wb["Periodo"], periodos_out)
    escribir_hoja(wb["Periodizacion"], periodiz_out)

    if proyecto_title != PLACEHOLDER_PRY:
        reemplazar_placeholder(wb, proyecto_title)

    out = BytesIO()
    wb.save(out)
    out.seek(0)
    return out, stats


if __name__ == "__main__":
    parser = argparse.ArgumentParser(
        description="Migra origen.xlsx → destino.xlsx (plantilla EIT App)"
    )
    parser.add_argument("--origen", default="origen.xlsx")
    parser.add_argument("--destino", default="destino.xlsx")
    parser.add_argument("--salida", default="output.xlsx")
    parser.add_argument(
        "--FP",
        default=DEFAULT_FECHA,
        metavar="DD/MM/AAAA",
        help="Fecha de conteo para columna 'fecha' en Periodizacion",
    )
    args = parser.parse_args()
    run(args.origen, args.destino, args.salida, args.FP)
