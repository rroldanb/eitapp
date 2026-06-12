from openpyxl import load_workbook
from datetime import date, time, datetime
from decimal import Decimal

from django.db import transaction

from apps.mandantes.models import Mandante, Contacto
from apps.proyectos.models import Proyecto
from apps.red_vial.models import (
    Calle, Nodo, Arco, Regulacion, CoeficienteCruce,
    Periodo, PuntoControl, Periodizacion, ParametroArco,
    FaseSemaforica, ConfiguracionTransyt,
)

SHEET_ORDER = [
    'Mandante', 'Contacto', 'Proyecto', 'Calle', 'Regulacion',
    'CoeficienteCruce', 'Periodo', 'ConfiguracionTransyt',
    'Nodo', 'Arco', 'PuntoControl', 'Periodizacion',
    'ParametroArco', 'FaseSemaforica',
]

SHEET_MODELS = {
    'Mandante': Mandante, 'Contacto': Contacto, 'Proyecto': Proyecto,
    'Calle': Calle, 'Nodo': Nodo, 'Arco': Arco, 'Regulacion': Regulacion,
    'CoeficienteCruce': CoeficienteCruce, 'Periodo': Periodo,
    'PuntoControl': PuntoControl, 'Periodizacion': Periodizacion,
    'ParametroArco': ParametroArco, 'FaseSemaforica': FaseSemaforica,
    'ConfiguracionTransyt': ConfiguracionTransyt,
}

DUPLICATE_KEYS = {
    'Mandante': ['name'],
    'Contacto': ['name', 'mandante'],
    'Proyecto': ['title'],
    'Calle': ['numero', 'proyecto_id'],
    'Nodo': ['numero', 'proyecto_id'],
    'Arco': ['nodo_origen', 'nodo_destino', 'proyecto_id'],
    'Regulacion': ['codigo'],
    'CoeficienteCruce': ['nomenclatura'],
    'Periodo': ['codigo', 'proyecto_id'],
    'PuntoControl': ['nodo', 'movimiento', 'proyecto_id'],
    'ParametroArco': ['punto_control'],
    'Periodizacion': ['fecha', 'pc_mov', 'hora'],
    'FaseSemaforica': ['punto_control', 'fase_numero'],
    'ConfiguracionTransyt': ['proyecto'],
}

REQUIRED_FIELDS = {
    'Mandante': ['name', 'location'],
    'Contacto': ['name', 'mandante'],
    'Proyecto': ['title', 'date_started', 'mandante'],
    'Calle': ['nombre', 'numero', 'proyecto'],
    'Nodo': ['numero', 'proyecto'],
    'Arco': ['nodo_origen', 'nodo_destino', 'longitud', 'proyecto'],
    'Regulacion': ['codigo', 'descripcion'],
    'CoeficienteCruce': ['nomenclatura', 'tipo_transporte', 'coeficiente', 'is_standard'],
    'Periodo': ['codigo', 'es_laboral', 'proyecto'],
    'PuntoControl': ['nodo', 'movimiento', 'is_prioritario', 'proyecto'],
    'Periodizacion': ['fecha', 'hora', 'pc', 'periodo', 'vl', 'txc', 'txb', 'c2e', 'c_mas2e', 'peat', 'cicl', 'moto'],
    'ParametroArco': ['punto_control', 'flujo_saturacion', 'ponderador_demora', 'ponderador_detencion', 'tiene_tarjeta_38', 'proyecto'],
    'FaseSemaforica': ['punto_control', 'fase_numero', 'verde_inicio', 'verde_fin', 'proyecto'],
    'ConfiguracionTransyt': ['proyecto', 'ciclo', 'W', 'K', 'perdida_inicial', 'ganancia_final'],
}

FK_FIELDS = {
    'Contacto': {'mandante': ('Mandante', 'name')},
    'Proyecto': {'mandante': ('Mandante', 'name')},
    'Calle': {'proyecto': ('Proyecto', 'title')},
    'Nodo': {'calle_1': ('Calle', 'nombre'), 'calle_2': ('Calle', 'nombre'), 'proyecto': ('Proyecto', 'title')},
    'Arco': {'nodo_origen': ('Nodo', 'numero', 'proyecto_id'), 'nodo_destino': ('Nodo', 'numero', 'proyecto_id'), 'proyecto': ('Proyecto', 'title')},
    'CoeficienteCruce': {'proyecto': ('Proyecto', 'title')},
    'Periodo': {'proyecto': ('Proyecto', 'title')},
    'PuntoControl': {'nodo': ('Nodo', 'numero', 'proyecto_id'), 'regulacion': ('Regulacion', 'codigo'), 'proyecto': ('Proyecto', 'title')},
    'Periodizacion': {'pc': ('PuntoControl', 'nombre', 'proyecto_id'), 'periodo': ('Periodo', 'codigo', 'proyecto_id'), 'proyecto': ('Proyecto', 'title')},
    'ParametroArco': {'punto_control': ('PuntoControl', 'nombre', 'proyecto_id'), 'proyecto': ('Proyecto', 'title')},
    'FaseSemaforica': {'punto_control': ('PuntoControl', 'nombre', 'proyecto_id'), 'proyecto': ('Proyecto', 'title')},
    'ConfiguracionTransyt': {'proyecto': ('Proyecto', 'title')},
}


def _choose(val, default=''):
    return val if val is not None else default


def _parse_date(val):
    if isinstance(val, date):
        return val
    if isinstance(val, str):
        val = val.strip()
        for fmt in ('%d/%m/%Y', '%Y-%m-%d', '%d-%m-%Y'):
            try:
                return datetime.strptime(val, fmt).date()
            except ValueError:
                pass
    return None


def _parse_time(val):
    if isinstance(val, time):
        return val
    if isinstance(val, datetime):
        return val.time()
    if isinstance(val, str):
        val = val.strip()
        for fmt in ('%H:%M', '%H:%M:%S'):
            try:
                return datetime.strptime(val, fmt).time()
            except ValueError:
                pass
    return None


def _parse_float(val):
    if val is None:
        return None
    if isinstance(val, (int, float, Decimal)):
        return float(val)
    if isinstance(val, str):
        val = val.strip().replace(',', '.')
        try:
            return float(val)
        except ValueError:
            pass
    return None


def _parse_int(val):
    if val is None:
        return None
    if isinstance(val, (int, float)):
        return int(val)
    if isinstance(val, str):
        val = val.strip()
        try:
            return int(val)
        except ValueError:
            pass
    return None


def _parse_bool(val):
    if isinstance(val, bool):
        return val
    if isinstance(val, str):
        return val.strip().upper() in ('SI', 'YES', 'TRUE', '1', 'S')
    if isinstance(val, (int, float)):
        return val == 1
    return False


def _trim_row(row):
    return [c.value if c.value is not None else '' for c in row]


def _header_index(headers, name):
    for i, h in enumerate(headers):
        if h and h.strip().lower() == name.strip().lower():
            return i
    return None


def parse_excel(file):
    wb = load_workbook(file, read_only=True, data_only=True)
    result = {}
    for ws in wb.worksheets:
        sheet_name = ws.title.strip()
        if sheet_name == "\U0001f4d6 README":
            continue
        rows_iter = ws.iter_rows(values_only=False)
        header_row = next(rows_iter, None)
        if header_row is None:
            continue
        headers = [h.value.strip() if h.value else '' for h in header_row]
        # skip type row (row 2) and description row (row 3) from template
        next(rows_iter, None)
        next(rows_iter, None)
        data_rows = []
        for row in rows_iter:
            vals = [c.value for c in row]
            if all(v is None or v == '' for v in vals):
                break
            row_dict = {}
            for i, val in enumerate(vals):
                if i < len(headers) and headers[i]:
                    row_dict[headers[i]] = val if val is not None else ''
            if row_dict:
                data_rows.append(row_dict)
        if data_rows:
            result[sheet_name] = data_rows
    wb.close()
    return result


class _VirtualObj:
    """minimal mock for FK resolution across sheets during validation.
    Stores the original lookup value for later re-resolution during execution.
    """
    def __init__(self, pk=0, value=''):
        self.pk = pk
        self.id = pk
        self._value = value


def validate_sheet(sheet_name, rows, proyecto=None, shared_cache=None):
    required = REQUIRED_FIELDS.get(sheet_name, [])
    fk_map = FK_FIELDS.get(sheet_name, {})
    valid = []
    errors = []
    ctx = {
        'proyecto': proyecto,
        'resolved': {},
        'shared_cache': shared_cache or {},
    }

    for idx, raw in enumerate(rows, 1):
        row_errors = []
        row = {}

        for key, val in raw.items():
            k = key.strip().lower().replace(' ', '_').replace('-', '_')
            row[k] = val

        row_ci = {k.lower(): v for k, v in row.items()}
        for f in required:
            val = row_ci.get(f.lower(), '')
            if val is None or val == '':
                row_errors.append(f"'{f}' es requerido y está vacío")

        if row_errors:
            errors.append({'row': idx, 'data': raw, 'errors': row_errors})
            continue

        parsed = _parse_row(sheet_name, row, idx, ctx, raw)
        if parsed['errors']:
            errors.append({'row': idx, 'data': raw, 'errors': parsed['errors']})
        else:
            valid.append(parsed['data'])

    # detect duplicates
    model = SHEET_MODELS.get(sheet_name)
    dup_keys = DUPLICATE_KEYS.get(sheet_name, [])
    duplicates = []
    new_rows = []
    for row_data in valid:
        filters = {}
        for k in dup_keys:
            v = row_data.get(k)
            if v is None and k == 'proyecto_id':
                v = row_data.get('proyecto')
                if v is not None and hasattr(v, 'id'):
                    v = v.id
            if v is not None:
                filters[k] = v
        if filters and model:
            try:
                if model.objects.filter(**filters).exists():
                    duplicates.append(row_data)
                    continue
            except Exception:
                pass
        new_rows.append(row_data)

    return {'sheet': sheet_name, 'valid': new_rows, 'duplicates': duplicates, 'errors': errors, 'total': len(rows)}


def _parse_row(sheet_name, row, idx, ctx, raw):
    errors = []
    data = {}

    if sheet_name == 'Mandante':
        data = {'name': row.get('name', ''), 'location': row.get('location', ''), 'details': _choose(row.get('details'))}

    elif sheet_name == 'Contacto':
        data = {
            'name': row.get('name', ''),
            'email': _choose(row.get('email')),
            'phone': _choose(row.get('phone')),
            'cargo': _choose(row.get('cargo')),
            'position': _choose(row.get('position')),
            'details': _choose(row.get('details')),
        }
        m = _resolve_fk('Contacto', 'mandante', row.get('mandante', ''), ctx)
        if m is None:
            errors.append(f"Mandante '{row.get('mandante', '')}' no encontrado")
        else:
            data['mandante'] = m

    elif sheet_name == 'Proyecto':
        title = row.get('title', '')
        d = _parse_date(row.get('date_started', ''))
        if d is None:
            errors.append("'date_started' no es una fecha válida")
        data = {'title': title, 'description': _choose(row.get('description')), 'date_started': d}
        m = _resolve_fk('Proyecto', 'mandante', row.get('mandante', ''), ctx)
        if m is None:
            errors.append(f"Mandante '{row.get('mandante', '')}' no encontrado")
        else:
            data['mandante'] = m

    elif sheet_name == 'Calle':
        data = {
            'nombre': row.get('nombre', '') or row.get('name', ''),
            'numero': _parse_int(row.get('numero')),
        }
        if data['numero'] is None:
            errors.append("'numero' debe ser un número entero")
        p = _resolve_fk('Calle', 'proyecto', row.get('proyecto', ''), ctx)
        if p is None:
            errors.append(f"Proyecto '{row.get('proyecto', '')}' no encontrado")
        else:
            data['proyecto'] = p

    elif sheet_name == 'Nodo':
        data = {
            'numero': _parse_int(row.get('numero')),
            'interseccion': _choose(row.get('interseccion')),
            'numero_pc': _parse_int(row.get('numero_pc')),
            'plano': _choose(row.get('plano')),
            'imagen': _choose(row.get('imagen')),
        }
        if data['numero'] is None:
            errors.append("'numero' debe ser un número entero")
        for fk_field, col_name in [('calle_1', 'calle_1'), ('calle_2', 'calle_2')]:
            val = row.get(col_name, '') or row.get(col_name.replace('_', '-'), '')
            if val:
                ref = _resolve_fk('Nodo', fk_field, val, ctx)
                if ref is None:
                    errors.append(f"Calle '{val}' no encontrada")
                else:
                    data[fk_field] = ref
        p = _resolve_fk('Nodo', 'proyecto', row.get('proyecto', ''), ctx)
        if p is None:
            errors.append(f"Proyecto '{row.get('proyecto', '')}' no encontrado")
        else:
            data['proyecto'] = p

    elif sheet_name == 'Arco':
        data = {'longitud': _parse_float(row.get('longitud'))}
        if data['longitud'] is None:
            errors.append("'longitud' debe ser un número")
        for fk_field, col in [('nodo_origen', 'nodo_origen'), ('nodo_destino', 'nodo_destino')]:
            val = row.get(col, '')
            ref = _resolve_fk('Arco', fk_field, val, ctx)
            if ref is None:
                errors.append(f"Nodo '{val}' no encontrado en este proyecto")
            else:
                data[fk_field] = ref
        p = _resolve_fk('Arco', 'proyecto', row.get('proyecto', ''), ctx)
        if p is None:
            errors.append(f"Proyecto '{row.get('proyecto', '')}' no encontrado")
        else:
            data['proyecto'] = p

    elif sheet_name == 'Regulacion':
        data = {'codigo': row.get('codigo', ''), 'descripcion': row.get('descripcion', '')}

    elif sheet_name == 'CoeficienteCruce':
        is_std = _parse_bool(row.get('is_standard', 'NO'))
        data = {
            'nomenclatura': row.get('nomenclatura', ''),
            'tipo_transporte': row.get('tipo_transporte', ''),
            'coeficiente': _parse_float(row.get('coeficiente')),
            'is_standard': is_std,
        }
        if data['coeficiente'] is None:
            errors.append("'coeficiente' debe ser un número")
        if not is_std:
            p = _resolve_fk('CoeficienteCruce', 'proyecto', row.get('proyecto', ''), ctx)
            if p is None:
                errors.append(f"Proyecto '{row.get('proyecto', '')}' no encontrado")
            else:
                data['proyecto'] = p

    elif sheet_name == 'Periodo':
        data = {
            'codigo': row.get('codigo', ''),
            'hora_inicio': _parse_time(row.get('hora_inicio')),
            'hora_fin': _parse_time(row.get('hora_fin')),
            'es_laboral': _parse_bool(row.get('es_laboral')),
        }
        p = _resolve_fk('Periodo', 'proyecto', row.get('proyecto', ''), ctx)
        if p is None:
            errors.append(f"Proyecto '{row.get('proyecto', '')}' no encontrado")
        else:
            data['proyecto'] = p

    elif sheet_name == 'PuntoControl':
        data = {
            'movimiento': _choose(row.get('movimiento')),
            'viraje': _choose(row.get('viraje')),
            'is_prioritario': _parse_bool(row.get('is_prioritario')),
            'numero_pistas': _parse_float(row.get('numero_pistas')),
        }
        nodo_ref = _resolve_fk('PuntoControl', 'nodo', row.get('nodo', ''), ctx)
        if nodo_ref is None:
            errors.append(f"Nodo '{row.get('nodo', '')}' no encontrado")
        else:
            data['nodo'] = nodo_ref
            if isinstance(nodo_ref, _VirtualObj):
                try:
                    nodo_num = int(nodo_ref._value)
                    data['nombre'] = f"PC-{nodo_num:02d}"
                except (ValueError, TypeError):
                    data['nombre'] = f"Nodo-{nodo_ref._value}"
            elif hasattr(nodo_ref, 'numero_pc') and nodo_ref.numero_pc:
                data['nombre'] = f"PC-{nodo_ref.numero_pc:02d}"
            elif hasattr(nodo_ref, 'numero'):
                data['nombre'] = f"Nodo-{nodo_ref.numero}"
        for fk_field, col in [('arco_entrada', 'arco_entrada'), ('arco_salida', 'arco_salida')]:
            val = row.get(col, '')
            if val:
                ref = _resolve_arco(val, ctx.get('proyecto'), ctx.get('shared_cache'))
                if ref is None:
                    errors.append(f"Arco '{val}' no encontrado")
                else:
                    data[fk_field] = ref
        reg = _resolve_fk('PuntoControl', 'regulacion', row.get('regulacion', ''), ctx)
        if row.get('regulacion') and reg is None:
            errors.append(f"Regulación '{row.get('regulacion', '')}' no encontrada")
        elif row.get('regulacion'):
            data['regulacion'] = reg
        p = _resolve_fk('PuntoControl', 'proyecto', row.get('proyecto', ''), ctx)
        if p is None:
            errors.append(f"Proyecto '{row.get('proyecto', '')}' no encontrado")
        else:
            data['proyecto'] = p

    elif sheet_name == 'Periodizacion':
        data = {
            'fecha': _parse_date(row.get('fecha')),
            'hora': _parse_time(row.get('hora')),
            'vl': _parse_int(row.get('vl')) or 0,
            'txc': _parse_int(row.get('txc')) or 0,
            'txb': _parse_int(row.get('txb')) or 0,
            'c2e': _parse_int(row.get('c2e')) or 0,
            'c_mas2e': _parse_int(row.get('c_mas2e')) or 0,
            'peat': _parse_int(row.get('peat')) or 0,
            'cicl': _parse_int(row.get('cicl')) or 0,
            'moto': _parse_int(row.get('moto')) or 0,
        }
        if data['fecha'] is None:
            errors.append("'fecha' no es una fecha válida")
        if data['hora'] is None:
            errors.append("'hora' no es una hora válida")
        pc_id_val = str(row.get('pc', '') or row.get('pc_mov', '') or '').strip()
        pc_ref = _resolve_fk('Periodizacion', 'pc', pc_id_val, ctx)
        if pc_ref is None:
            errors.append(f"PuntoControl '{pc_id_val}' no encontrado")
        else:
            data['pc'] = pc_ref
            # Auto-generate pc_mov from PuntoControl's arc codigos
            if isinstance(pc_ref, _VirtualObj):
                # PuntoControl comes from the same import, can't verify arcs now;
                # _execute_sheet will recompute pc_mov from the real PC object
                data['pc_mov'] = pc_id_val
            else:
                arco_s = getattr(pc_ref, 'arco_salida', None)
                arco_e = getattr(pc_ref, 'arco_entrada', None)
                if arco_s is not None and arco_e is not None:
                    data['pc_mov'] = f"{arco_s.codigo_arco}_{arco_e.codigo_arco}"
                else:
                    errors.append(f"PuntoControl '{pc_id_val}' no tiene arcos de entrada/salida definidos")
        per_ref = _resolve_fk('Periodizacion', 'periodo', row.get('periodo', ''), ctx)
        if per_ref is None:
            errors.append(f"Periodo '{row.get('periodo', '')}' no encontrado")
        else:
            data['periodo'] = per_ref
        p = _resolve_fk('Periodizacion', 'proyecto', row.get('proyecto', ''), ctx)
        if p is None:
            errors.append(f"Proyecto '{row.get('proyecto', '')}' no encontrado")
        else:
            data['proyecto'] = p

    elif sheet_name == 'ParametroArco':
        data = {
            'flujo_saturacion': _parse_float(row.get('flujo_saturacion')) or 1800.0,
            'ponderador_demora': _parse_float(row.get('ponderador_demora')) or 1.0,
            'ponderador_detencion': _parse_float(row.get('ponderador_detencion')) or 1.0,
            'capacidad_cola': _parse_float(row.get('capacidad_cola')),
            'tiene_tarjeta_38': _parse_bool(row.get('tiene_tarjeta_38')),
        }
        pc_ref = _resolve_fk('ParametroArco', 'punto_control', row.get('punto_control', ''), ctx)
        if pc_ref is None:
            errors.append(f"PuntoControl '{row.get('punto_control', '')}' no encontrado")
        else:
            data['punto_control'] = pc_ref
        p = _resolve_fk('ParametroArco', 'proyecto', row.get('proyecto', ''), ctx)
        if p is None:
            errors.append(f"Proyecto '{row.get('proyecto', '')}' no encontrado")
        else:
            data['proyecto'] = p

    elif sheet_name == 'FaseSemaforica':
        data = {
            'fase_numero': _parse_int(row.get('fase_numero')),
            'verde_inicio': _parse_float(row.get('verde_inicio')) or 0.0,
            'verde_fin': _parse_float(row.get('verde_fin')) or 0.0,
        }
        if data['fase_numero'] is None:
            errors.append("'fase_numero' debe ser un número entero")
        pc_ref = _resolve_fk('FaseSemaforica', 'punto_control', row.get('punto_control', ''), ctx)
        if pc_ref is None:
            errors.append(f"PuntoControl '{row.get('punto_control', '')}' no encontrado")
        else:
            data['punto_control'] = pc_ref
        p = _resolve_fk('FaseSemaforica', 'proyecto', row.get('proyecto', ''), ctx)
        if p is None:
            errors.append(f"Proyecto '{row.get('proyecto', '')}' no encontrado")
        else:
            data['proyecto'] = p

    elif sheet_name == 'ConfiguracionTransyt':
        data = {
            'ciclo': _parse_int(row.get('ciclo')) or 60,
            'W': _parse_float(row.get('W') or row.get('w')) or 10.0,
            'K': _parse_float(row.get('K') or row.get('k')) or 0.5,
            'perdida_inicial': _parse_float(row.get('perdida_inicial')) or 2.0,
            'ganancia_final': _parse_float(row.get('ganancia_final')) or 1.0,
        }
        p = _resolve_fk('ConfiguracionTransyt', 'proyecto', row.get('proyecto', ''), ctx)
        if p is None:
            errors.append(f"Proyecto '{row.get('proyecto', '')}' no encontrado")
        else:
            data['proyecto'] = p

    return {'data': data, 'errors': errors}


def _resolve_puntocontrol_por_nombre(value, ctx):
    """Resolve PuntoControl by nombre (PC-XX or Nodo-XX), which is a @property."""
    proyecto = ctx.get('proyecto')
    if not proyecto:
        return None
    value = str(value).strip()
    # Check shared_cache first (cross-sheet validation)
    sc = ctx.get('shared_cache', {})
    sc_key = ('PuntoControl', 'nombre', value, str(proyecto.id))
    if sc.get(sc_key):
        return _VirtualObj(pk=0, value=value)
    qs = PuntoControl.objects.filter(proyecto=proyecto)
    if value.startswith('PC-'):
        try:
            num = int(value[3:])
            return qs.filter(nodo__numero_pc=num).first()
        except ValueError:
            return None
    elif value.startswith('Nodo-') or value.startswith('Nodo'):
        num_str = value.replace('Nodo-', '').replace('Nodo', '')
        try:
            num = int(num_str)
            return qs.filter(nodo__numero=num).first()
        except ValueError:
            return None
    return None


def _resolve_fk(sheet_name, field, value, ctx):
    if not value:
        return None
    value = str(value).strip()
    fk_info = FK_FIELDS.get(sheet_name, {}).get(field)
    if not fk_info:
        return None
    model_name = fk_info[0]
    lookup_field = fk_info[1]
    has_proyecto = len(fk_info) > 2
    cache_key = f'{model_name}:{value}'
    if has_proyecto and ctx.get('proyecto'):
        cache_key = f'{model_name}:{ctx["proyecto"].id}:{value}'

    if cache_key in ctx['resolved']:
        return ctx['resolved'][cache_key]

    model = _get_model(model_name)
    if model is None:
        return None

    # Special case: PuntoControl.nombre is a @property, not a DB field
    if model_name == 'PuntoControl' and lookup_field == 'nombre':
        resolved = _resolve_puntocontrol_por_nombre(value, ctx)
        if resolved:
            ctx['resolved'][cache_key] = resolved
            return resolved
        return None

    filters = {lookup_field: value}
    if has_proyecto and ctx.get('proyecto'):
        filters['proyecto_id'] = ctx['proyecto'].id
    elif has_proyecto and ctx.get('proyecto_id'):
        filters['proyecto_id'] = ctx['proyecto_id']

    try:
        objs = model.objects.filter(**filters)
        obj = objs.first()
        if obj:
            ctx['resolved'][cache_key] = obj
            return obj
    except Exception:
        pass

    # check shared_cache (cross-sheet validation)
    sc = ctx.get('shared_cache', {})
    sc_key = (model_name, lookup_field, value)
    if has_proyecto and ctx.get('proyecto'):
        sc_key = sc_key + (str(ctx['proyecto'].id),)
    if sc.get(sc_key):
        mock = _VirtualObj(pk=0, value=value)
        ctx['resolved'][cache_key] = mock
        return mock

    return None


def _resolve_arco(val, proyecto, shared_cache=None):
    if not val or not proyecto:
        return None
    val = str(val).strip()
    parts = val.replace('>', '>').split('>')
    if len(parts) != 2:
        return None
    try:
        origen = int(parts[0].strip())
        destino = int(parts[1].strip())
    except ValueError:
        return None
    # Check shared_cache first (cross-sheet validation)
    if shared_cache:
        sc_key = ('Arco', '_key', f'{origen}>{destino}', str(proyecto.id))
        if sc_key in shared_cache:
            return _VirtualObj(pk=0, value=f'{origen}>{destino}')
    try:
        return Arco.objects.get(
            nodo_origen__numero=origen,
            nodo_destino__numero=destino,
            proyecto_id=proyecto.id,
        )
    except Arco.DoesNotExist:
        return None


def _resolve_sentinel_arco(proyecto):
    """Get or create "1>1" sentinel Arco for fallback when PC arco is missing."""
    if not proyecto:
        return None
    try:
        nodo1, _ = Nodo.objects.get_or_create(
            numero=1, proyecto=proyecto,
            defaults={'interseccion': None, 'calle_1': None, 'calle_2': None},
        )
        arco, _ = Arco.objects.get_or_create(
            nodo_origen=nodo1, nodo_destino=nodo1, proyecto=proyecto,
            defaults={'longitud': 1},
        )
        return arco
    except Exception:
        return None


def _get_model(name):
    models = {
        'Mandante': Mandante,
        'Contacto': Contacto,
        'Proyecto': Proyecto,
        'Calle': Calle,
        'Nodo': Nodo,
        'Arco': Arco,
        'Regulacion': Regulacion,
        'CoeficienteCruce': CoeficienteCruce,
        'Periodo': Periodo,
        'PuntoControl': PuntoControl,
        'Periodizacion': Periodizacion,
        'ParametroArco': ParametroArco,
        'FaseSemaforica': FaseSemaforica,
        'ConfiguracionTransyt': ConfiguracionTransyt,
    }
    return models.get(name)


def validate_selection(parsed_data, selected_sheets, proyecto=None):
    sheets_in_order = [s for s in SHEET_ORDER if s in selected_sheets]
    results = {}
    total_valid = 0
    total_errors = 0
    total_duplicates = 0
    shared_cache = {}
    for sheet_name in sheets_in_order:
        rows = parsed_data.get(sheet_name, [])
        if not rows:
            continue
        res = validate_sheet(sheet_name, rows, proyecto, shared_cache)
        # clean mock objects + date/time/Decimal from valid/duplicate rows for session storage
        for row_data in res['valid']:
            _restore_fk_values(sheet_name, row_data)
            _sanitize_for_session(row_data)
        for row_data in res.get('duplicates', []):
            _restore_fk_values(sheet_name, row_data)
            _sanitize_for_session(row_data)
        results[sheet_name] = res
        # register valid rows in shared_cache so later sheets resolve FKs to them
        for other_sheet, fk_dict in FK_FIELDS.items():
            for field_name, fk_info in fk_dict.items():
                if fk_info[0] == sheet_name:
                    lookup_field = fk_info[1]
                    for row_data in res['valid']:
                        val = row_data.get(lookup_field)
                        if val is not None:
                            cache_key = (sheet_name, lookup_field, str(val))
                            if len(fk_info) > 2 and proyecto:
                                cache_key = cache_key + (str(proyecto.id),)
                            shared_cache[cache_key] = True
        # Register Arco rows with composite key (origen>destino) for PuntoControl cross-sheet resolution
        if sheet_name == 'Arco':
            for row_data in res['valid']:
                origen = row_data.get('nodo_origen')
                destino = row_data.get('nodo_destino')
                if origen is not None and destino is not None:
                    arco_key = f'{origen}>{destino}'
                    sc_key = ('Arco', '_key', arco_key, str(proyecto.id)) if proyecto else ('Arco', '_key', arco_key)
                    shared_cache[sc_key] = True
        total_valid += len(res['valid'])
        total_errors += len(res['errors'])
        total_duplicates += len(res.get('duplicates', []))
    return {'results': results, 'total_valid': total_valid, 'total_errors': total_errors, 'total_duplicates': total_duplicates}


def _restore_fk_values(sheet_name, data):
    """replace FK values (mock or real model instances) with raw strings for session storage"""
    for field, fk_info in FK_FIELDS.get(sheet_name, {}).items():
        val = data.get(field)
        if isinstance(val, _VirtualObj):
            data[field] = val._value
        elif hasattr(val, '_meta'):  # real Django model instance
            lookup_field = fk_info[1]
            lookup_val = getattr(val, lookup_field, None)
            data[field] = str(lookup_val) if lookup_val is not None else ''
    # Handle arco_entrada/arco_salida custom fields for PuntoControl
    if sheet_name == 'PuntoControl':
        for field in ('arco_entrada', 'arco_salida'):
            val = data.get(field)
            if isinstance(val, _VirtualObj):
                data[field] = val._value
            elif hasattr(val, '_meta') and hasattr(val, 'nodo_origen'):  # real Arco instance
                data[field] = f"{val.nodo_origen.numero}>{val.nodo_destino.numero}"
    return data


def _sanitize_for_session(data):
    """convert date/time/Decimal objects to JSON-safe strings"""
    for k, v in data.items():
        if isinstance(v, (date, time)):
            data[k] = v.isoformat()
        elif isinstance(v, Decimal):
            data[k] = float(v)
    return data


@transaction.atomic
def execute_import(validation_result, proyecto, user, update_duplicates=None):
    results = validation_result['results']
    report = {}

    for sheet_name in SHEET_ORDER:
        if sheet_name not in results:
            continue
        res = results[sheet_name]

        # re-resolve FK values from DB (earlier sheets were already executed)
        fk_ctx = {'proyecto': proyecto, 'resolved': {}}
        for row_data in res['valid'] + res.get('duplicates', []):
            row_proyecto = row_data.get('proyecto')
            if isinstance(row_proyecto, Proyecto):
                row_fk_ctx = {'proyecto': row_proyecto, 'resolved': {}}
            elif isinstance(row_proyecto, str):
                p = _resolve_fk(sheet_name, 'proyecto', row_proyecto, fk_ctx)
                if isinstance(p, Proyecto):
                    row_data['proyecto'] = p
                    row_fk_ctx = {'proyecto': p, 'resolved': {}}
                else:
                    row_fk_ctx = fk_ctx
            else:
                row_fk_ctx = fk_ctx
            for field in FK_FIELDS.get(sheet_name, {}):
                val = row_data.get(field)
                if val and isinstance(val, str):
                    resolved = _resolve_fk(sheet_name, field, val, row_fk_ctx)
                    if resolved and not isinstance(resolved, _VirtualObj):
                        row_data[field] = resolved
        # Re-resolve arco_entrada/arco_salida for PuntoControl (custom resolver, not in FK_FIELDS)
        if sheet_name == 'PuntoControl':
            for row_data in res['valid'] + res.get('duplicates', []):
                row_proyecto = row_data.get('proyecto')
                if isinstance(row_proyecto, Proyecto):
                    arco_ctx = row_proyecto
                else:
                    arco_ctx = proyecto
                for field in ('arco_entrada', 'arco_salida'):
                    val = row_data.get(field)
                    if isinstance(val, _VirtualObj):
                        val = val._value
                    if val and isinstance(val, str):
                        resolved = _resolve_arco(val, arco_ctx)
                        if resolved and not isinstance(resolved, _VirtualObj):
                            row_data[field] = resolved
        dups = res.get('duplicates', [])
        should_update = update_duplicates and update_duplicates.get(sheet_name, True)
        sheet_report = {'inserted': 0, 'updated': 0, 'rejected': [], 'skipped_duplicates': 0, 'valid_count': len(res['valid']), 'duplicate_count': len(dups), 'error_count': len(res['errors'])}
        for error_row in res['errors']:
            sheet_report['rejected'].append({'row': error_row['row'], 'reason': '; '.join(error_row['errors'])})

        _execute_sheet(sheet_name, res['valid'], proyecto, user, sheet_report)

        if should_update:
            _execute_sheet(sheet_name, dups, proyecto, user, sheet_report)
        else:
            sheet_report['skipped_duplicates'] = len(dups)

        report[sheet_name] = sheet_report

    return report


def _execute_sheet(sheet_name, rows, proyecto, user, sheet_report):
    for data in rows:
        try:
            with transaction.atomic():
                if sheet_name == 'Mandante':
                    obj, created = Mandante.objects.get_or_create(name=data['name'], defaults={'location': data['location'], 'details': data['details']})
                    if not created:
                        obj.location = data['location']
                        obj.details = data['details']
                        obj.save()
                    if created:
                        sheet_report['inserted'] += 1
                    else:
                        sheet_report['updated'] += 1

                elif sheet_name == 'Contacto':
                    obj, created = Contacto.objects.get_or_create(
                        name=data['name'], mandante=data['mandante'],
                        defaults={'email': data.get('email', ''), 'phone': data.get('phone', ''), 'cargo': data.get('cargo', ''), 'position': data.get('position', ''), 'details': data.get('details', '')}
                    )
                    if not created:
                        for k in ('email', 'phone', 'cargo', 'position', 'details'):
                            setattr(obj, k, data.get(k, ''))
                        obj.mandante = data['mandante']
                        obj.save()
                    if created:
                        sheet_report['inserted'] += 1
                    else:
                        sheet_report['updated'] += 1

                elif sheet_name == 'Proyecto':
                    obj, created = Proyecto.objects.get_or_create(
                        title=data['title'],
                        defaults={'description': data.get('description', ''), 'date_started': data['date_started'], 'mandante': data['mandante'], 'user': user}
                    )
                    if not created:
                        obj.description = data.get('description', '')
                        obj.date_started = data['date_started']
                        obj.mandante = data['mandante']
                        obj.save()
                    if created:
                        sheet_report['inserted'] += 1
                    else:
                        sheet_report['updated'] += 1

                elif sheet_name == 'Calle':
                    try:
                        obj = Calle.objects.get(numero=data['numero'], proyecto=data['proyecto'])
                        obj.nombre = data['nombre']
                        obj.save()
                        sheet_report['updated'] += 1
                    except Calle.DoesNotExist:
                        Calle.objects.create(nombre=data['nombre'], numero=data['numero'], proyecto=data['proyecto'])
                        sheet_report['inserted'] += 1

                elif sheet_name == 'Nodo':
                    try:
                        obj = Nodo.objects.get(numero=data['numero'], proyecto=data['proyecto'])
                        for k in ('interseccion', 'numero_pc', 'plano', 'imagen', 'calle_1', 'calle_2'):
                            if k in data:
                                setattr(obj, k, data[k])
                        obj.save()
                        sheet_report['updated'] += 1
                    except Nodo.DoesNotExist:
                        Nodo.objects.create(proyecto=data.pop('proyecto'), **data)
                        sheet_report['inserted'] += 1

                elif sheet_name == 'Arco':
                    try:
                        obj = Arco.objects.get(nodo_origen=data['nodo_origen'], nodo_destino=data['nodo_destino'], proyecto=data['proyecto'])
                        obj.longitud = data['longitud']
                        obj.save()
                        sheet_report['updated'] += 1
                    except Arco.DoesNotExist:
                        Arco.objects.create(nodo_origen=data['nodo_origen'], nodo_destino=data['nodo_destino'], longitud=data['longitud'], proyecto=data['proyecto'])
                        sheet_report['inserted'] += 1

                elif sheet_name == 'Regulacion':
                    obj, created = Regulacion.objects.get_or_create(codigo=data['codigo'], defaults={'descripcion': data['descripcion']})
                    if not created:
                        obj.descripcion = data['descripcion']
                        obj.save()
                        sheet_report['updated'] += 1
                    else:
                        sheet_report['inserted'] += 1

                elif sheet_name == 'CoeficienteCruce':
                    filters = {'nomenclatura': data['nomenclatura']}
                    if data['is_standard']:
                        filters['proyecto__isnull'] = True
                    elif 'proyecto' in data:
                        filters['proyecto'] = data['proyecto']
                    try:
                        obj = CoeficienteCruce.objects.get(**filters)
                        obj.tipo_transporte = data['tipo_transporte']
                        obj.coeficiente = data['coeficiente']
                        obj.is_standard = data['is_standard']
                        obj.save()
                        sheet_report['updated'] += 1
                    except CoeficienteCruce.DoesNotExist:
                        CoeficienteCruce.objects.create(
                            nomenclatura=data['nomenclatura'], tipo_transporte=data['tipo_transporte'],
                            coeficiente=data['coeficiente'], is_standard=data['is_standard'],
                            proyecto=data.get('proyecto'),
                        )
                        sheet_report['inserted'] += 1

                elif sheet_name == 'Periodo':
                    try:
                        obj = Periodo.objects.get(codigo=data['codigo'], proyecto=data['proyecto'])
                        obj.hora_inicio = data.get('hora_inicio')
                        obj.hora_fin = data.get('hora_fin')
                        obj.es_laboral = data['es_laboral']
                        obj.save()
                        sheet_report['updated'] += 1
                    except Periodo.DoesNotExist:
                        Periodo.objects.create(codigo=data['codigo'], hora_inicio=data.get('hora_inicio'), hora_fin=data.get('hora_fin'), es_laboral=data['es_laboral'], proyecto=data['proyecto'])
                        sheet_report['inserted'] += 1

                elif sheet_name == 'PuntoControl':
                    data.pop('nombre', None)
                    for field in ('arco_entrada', 'arco_salida'):
                        if not isinstance(data.get(field), Arco):
                            sentinel = _resolve_sentinel_arco(data.get('proyecto'))
                            if sentinel:
                                data[field] = sentinel
                    try:
                        obj = PuntoControl.objects.get(nodo=data['nodo'], movimiento=data['movimiento'], proyecto=data['proyecto'])
                        for k in ('viraje', 'is_prioritario', 'arco_entrada', 'arco_salida', 'regulacion', 'numero_pistas'):
                            if k in data:
                                setattr(obj, k, data[k])
                        obj.save()
                        sheet_report['updated'] += 1
                    except PuntoControl.DoesNotExist:
                        PuntoControl.objects.create(proyecto=data.pop('proyecto'), **data)
                        sheet_report['inserted'] += 1

                elif sheet_name == 'Periodizacion':
                    data.pop('proyecto', None)
                    # Recompute pc_mov from re-resolved PuntoControl's arcs
                    pc = data.get('pc')
                    if isinstance(pc, PuntoControl) and pc.arco_salida_id and pc.arco_entrada_id:
                        pc_mov = f"{pc.arco_salida.codigo_arco}_{pc.arco_entrada.codigo_arco}"
                        data['pc_mov'] = pc_mov
                    elif 'pc_mov' not in data or not data['pc_mov']:
                        sheet_report['rejected'].append({'row': data.get('_original_row', '?'), 'reason': 'PuntoControl sin arcos de entrada/salida'})
                        continue
                    obj, created = Periodizacion.objects.get_or_create(
                        fecha=data['fecha'], pc_mov=data['pc_mov'], hora=data['hora'],
                        defaults={k: data.get(k) for k in ('pc', 'periodo', 'vl', 'txc', 'txb', 'c2e', 'c_mas2e', 'peat', 'cicl', 'moto')},
                    )
                    if not created:
                        for k in ('pc', 'periodo', 'vl', 'txc', 'txb', 'c2e', 'c_mas2e', 'peat', 'cicl', 'moto'):
                            setattr(obj, k, data.get(k))
                        obj.save()
                        sheet_report['updated'] += 1
                    else:
                        sheet_report['inserted'] += 1

                elif sheet_name == 'ParametroArco':
                    try:
                        obj = ParametroArco.objects.get(punto_control=data['punto_control'])
                        for k in ('flujo_saturacion', 'ponderador_demora', 'ponderador_detencion', 'capacidad_cola', 'tiene_tarjeta_38'):
                            if k in data:
                                setattr(obj, k, data[k])
                        obj.save()
                        sheet_report['updated'] += 1
                    except ParametroArco.DoesNotExist:
                        ParametroArco.objects.create(proyecto=data.pop('proyecto'), **data)
                        sheet_report['inserted'] += 1

                elif sheet_name == 'FaseSemaforica':
                    try:
                        obj = FaseSemaforica.objects.get(punto_control=data['punto_control'], fase_numero=data['fase_numero'])
                        obj.verde_inicio = data['verde_inicio']
                        obj.verde_fin = data['verde_fin']
                        obj.save()
                        sheet_report['updated'] += 1
                    except FaseSemaforica.DoesNotExist:
                        FaseSemaforica.objects.create(proyecto=data.pop('proyecto'), **data)
                        sheet_report['inserted'] += 1

                elif sheet_name == 'ConfiguracionTransyt':
                    obj, created = ConfiguracionTransyt.objects.get_or_create(
                        proyecto=data['proyecto'],
                        defaults={'ciclo': data['ciclo'], 'W': data['W'], 'K': data['K'], 'perdida_inicial': data['perdida_inicial'], 'ganancia_final': data['ganancia_final']}
                    )
                    if not created:
                        for k in ('ciclo', 'W', 'K', 'perdida_inicial', 'ganancia_final'):
                            setattr(obj, k, data[k])
                        obj.save()
                        sheet_report['updated'] += 1
                    else:
                        sheet_report['inserted'] += 1

        except Exception as e:
            sheet_report['rejected'].append({'row': data.get('_original_row', '?'), 'reason': str(e)})


def analyze_parsed_data(parsed_data):
    """Extract available project datasets from parsed Excel data."""
    proyectos = parsed_data.get('Proyecto', [])
    datasets = []
    for p in proyectos:
        title = (p.get('title') or '').strip()
        if title:
            datasets.append({
                'title': title,
                'mandante': (p.get('mandante') or '').strip(),
            })
    return datasets


def filter_by_dataset(parsed_data, dataset_title):
    """Filter parsed rows to only include those matching the selected project."""
    dataset_title = dataset_title.strip()
    filtered = {}
    for sheet_name, rows in parsed_data.items():
        if not rows:
            continue
        # Find which key in the row corresponds to 'proyecto'
        proyecto_key = None
        for key in rows[0]:
            if key.lower().replace(' ', '_').replace('-', '_') == 'proyecto':
                proyecto_key = key
                break
        if sheet_name == 'Proyecto':
            filtered[sheet_name] = [r for r in rows
                                    if (r.get('title') or '').strip() == dataset_title]
        elif proyecto_key:
            filtered[sheet_name] = [r for r in rows
                                    if (r.get(proyecto_key) or '').strip() == dataset_title]
        else:
            filtered[sheet_name] = rows
    return filtered


def reassign_to_project(parsed_data, proyecto_title):
    """Reassign all rows to point to a different project title."""
    proyecto_title = proyecto_title.strip()
    for sheet_name, rows in parsed_data.items():
        if not rows:
            continue
        if sheet_name == 'Proyecto':
            for r in rows:
                r['title'] = proyecto_title
            continue
        proyecto_key = None
        for key in rows[0]:
            if key.lower().replace(' ', '_').replace('-', '_') == 'proyecto':
                proyecto_key = key
                break
        if proyecto_key:
            for r in rows:
                r[proyecto_key] = proyecto_title
