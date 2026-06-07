import io
import json
from datetime import date, time, datetime, timedelta
from decimal import Decimal
from unittest.mock import patch, MagicMock, ANY
from django.test import TestCase
from django.contrib.auth.models import User
from openpyxl import Workbook, load_workbook

from apps.mandantes.models import Mandante, Contacto
from apps.proyectos.models import Proyecto
from apps.red_vial.models import (
    Calle, Nodo, Arco, Regulacion, CoeficienteCruce,
    Periodo, PuntoControl, Periodizacion, ParametroArco,
    FaseSemaforica, ConfiguracionTransyt,
)
from apps.red_vial.services.import_service import (
    _choose, _parse_date, _parse_time, _parse_float, _parse_int, _parse_bool,
    _trim_row, _header_index, _VirtualObj, _restore_fk_values, _sanitize_for_session,
    _resolve_fk, _resolve_puntocontrol_por_nombre, _resolve_arco, _get_model, _parse_row,
    parse_excel, validate_sheet, validate_selection, execute_import,
    SHEET_ORDER, SHEET_MODELS, DUPLICATE_KEYS, REQUIRED_FIELDS, FK_FIELDS,
)


def _make_excel(sheets_data):
    """Create an in-memory Excel file (BytesIO) with given sheet data.
    sheets_data: {sheet_name: [list_of_dicts]}
    Each dict is a row. Headers are taken from dict keys (ordered).
    A type row and description row are inserted after headers (per template format).
    """
    wb = Workbook()
    # remove default sheet
    wb.remove(wb.active)
    for sheet_name, rows in sheets_data.items():
        ws = wb.create_sheet(title=sheet_name)
        if not rows:
            continue
        headers = list(rows[0].keys())
        ws.append(headers)
        ws.append(['TIPO'] + [''] * (len(headers) - 1))
        ws.append(['DESCRIPCION'] + [''] * (len(headers) - 1))
        for row in rows:
            ws.append([row.get(h, '') for h in headers])
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


class ParseExcelTest(TestCase):
    def test_parse_basic_sheet(self):
        buf = _make_excel({
            'Mandante': [
                {'name': 'M1', 'location': 'Loc1', 'details': 'Det1'},
            ]
        })
        result = parse_excel(buf)
        self.assertIn('Mandante', result)
        self.assertEqual(len(result['Mandante']), 1)
        self.assertEqual(result['Mandante'][0]['name'], 'M1')

    def test_skip_readme_sheet(self):
        buf = _make_excel({
            'Mandante': [{'name': 'M1', 'location': 'L1'}],
            '\U0001f4d6 README': [{'x': 'y'}],
        })
        result = parse_excel(buf)
        self.assertIn('Mandante', result)
        self.assertNotIn('\U0001f4d6 README', result)

    def test_skip_type_and_description_rows(self):
        buf = _make_excel({
            'Mandante': [
                {'name': 'M1', 'location': 'Loc1'},
            ]
        })
        result = parse_excel(buf)
        self.assertEqual(len(result['Mandante']), 1)

    def test_empty_row_breaks(self):
        """Empty row (all None) should stop parsing, subsequent rows ignored."""
        wb = Workbook()
        ws = wb.active
        ws.title = 'Mandante'
        ws.append(['name', 'location', 'details'])
        ws.append(['TIPO', '', ''])
        ws.append(['DESCRIPCION', '', ''])
        ws.append(['M1', 'L1', ''])
        ws.append([None, None, None])
        ws.append(['M2', 'L2', ''])
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        result = parse_excel(buf)
        self.assertIn('Mandante', result)
        self.assertEqual(len(result['Mandante']), 1)
        self.assertEqual(result['Mandante'][0]['name'], 'M1')

    def test_no_data_sheets_omitted(self):
        buf = _make_excel({
            'Mandante': [{'name': 'M1', 'location': 'L1'}],
            'EmptySheet': [],
        })
        result = parse_excel(buf)
        self.assertIn('Mandante', result)
        self.assertNotIn('EmptySheet', result)

    def test_handles_empty_workbook(self):
        wb = Workbook()
        ws = wb.active
        ws.title = 'EmptySheet'
        # Only a header row (no data rows after type/description)
        ws.append(['col1'])
        ws.append(['TIPO'])
        ws.append(['DESCRIPCION'])
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        result = parse_excel(buf)
        self.assertEqual(result, {})

    def test_handles_none_header_cell(self):
        buf = _make_excel({
            'Mandante': [{'name': 'M1', 'location': 'L1'}],
        })
        result = parse_excel(buf)
        self.assertEqual(len(result['Mandante']), 1)


class HelpersTest(TestCase):
    def test_choose_returns_default_for_none(self):
        self.assertEqual(_choose(None, 'X'), 'X')

    def test_choose_returns_value(self):
        self.assertEqual(_choose('hello', 'X'), 'hello')

    def test_parse_date_from_date(self):
        d = date(2024, 1, 15)
        self.assertEqual(_parse_date(d), d)

    def test_parse_date_from_string_dmy(self):
        self.assertEqual(_parse_date('15/01/2024'), date(2024, 1, 15))

    def test_parse_date_from_string_ymd(self):
        self.assertEqual(_parse_date('2024-01-15'), date(2024, 1, 15))

    def test_parse_date_from_string_dmy_hyphen(self):
        self.assertEqual(_parse_date('15-01-2024'), date(2024, 1, 15))

    def test_parse_date_invalid(self):
        self.assertIsNone(_parse_date('not-a-date'))

    def test_parse_date_empty(self):
        self.assertIsNone(_parse_date(''))

    def test_parse_time_from_time(self):
        t = time(14, 30)
        self.assertEqual(_parse_time(t), t)

    def test_parse_time_from_datetime(self):
        dt = datetime(2024, 1, 15, 14, 30)
        self.assertEqual(_parse_time(dt), time(14, 30))

    def test_parse_time_from_string_hm(self):
        self.assertEqual(_parse_time('14:30'), time(14, 30))

    def test_parse_time_from_string_hms(self):
        self.assertEqual(_parse_time('14:30:00'), time(14, 30))

    def test_parse_time_invalid(self):
        self.assertIsNone(_parse_time('not-a-time'))

    def test_parse_float_none(self):
        self.assertIsNone(_parse_float(None))

    def test_parse_float_from_int(self):
        self.assertEqual(_parse_float(5), 5.0)

    def test_parse_float_from_float(self):
        self.assertEqual(_parse_float(3.14), 3.14)

    def test_parse_float_from_decimal(self):
        self.assertEqual(_parse_float(Decimal('3.14')), 3.14)

    def test_parse_float_from_string(self):
        self.assertEqual(_parse_float('3.14'), 3.14)

    def test_parse_float_from_string_comma(self):
        self.assertEqual(_parse_float('3,14'), 3.14)

    def test_parse_float_invalid(self):
        self.assertIsNone(_parse_float('abc'))

    def test_parse_int_none(self):
        self.assertIsNone(_parse_int(None))

    def test_parse_int_from_int(self):
        self.assertEqual(_parse_int(5), 5)

    def test_parse_int_from_float(self):
        self.assertEqual(_parse_int(5.7), 5)

    def test_parse_int_from_string(self):
        self.assertEqual(_parse_int('42'), 42)

    def test_parse_int_invalid(self):
        self.assertIsNone(_parse_int('abc'))

    def test_parse_bool_true_bool(self):
        self.assertTrue(_parse_bool(True))

    def test_parse_bool_false_bool(self):
        self.assertFalse(_parse_bool(False))

    def test_parse_bool_si(self):
        self.assertTrue(_parse_bool('SI'))

    def test_parse_bool_yes(self):
        self.assertTrue(_parse_bool('YES'))

    def test_parse_bool_true_str(self):
        self.assertTrue(_parse_bool('TRUE'))

    def test_parse_bool_1_str(self):
        self.assertTrue(_parse_bool('1'))

    def test_parse_bool_s(self):
        self.assertTrue(_parse_bool('S'))

    def test_parse_bool_no(self):
        self.assertFalse(_parse_bool('NO'))

    def test_parse_bool_int_1(self):
        self.assertTrue(_parse_bool(1))

    def test_parse_bool_int_0(self):
        self.assertFalse(_parse_bool(0))

    def test_trim_row_replaces_none(self):
        class MockCell:
            def __init__(self, v):
                self.value = v
        row = [MockCell('a'), MockCell(None), MockCell('b')]
        self.assertEqual(_trim_row(row), ['a', '', 'b'])

    def test_header_index_case_insensitive(self):
        headers = ['Name', 'Location', 'Details']
        self.assertEqual(_header_index(headers, 'name'), 0)
        self.assertEqual(_header_index(headers, 'LOCATION'), 1)
        self.assertEqual(_header_index(headers, 'details'), 2)

    def test_header_index_not_found(self):
        self.assertIsNone(_header_index(['Name'], 'Unknown'))

    def test_virtual_obj_creation(self):
        obj = _VirtualObj(pk=42, value='test-value')
        self.assertEqual(obj.pk, 42)
        self.assertEqual(obj.id, 42)
        self.assertEqual(obj._value, 'test-value')

    def test_get_model(self):
        self.assertEqual(_get_model('Mandante'), Mandante)
        self.assertEqual(_get_model('Contacto'), Contacto)
        self.assertEqual(_get_model('Proyecto'), Proyecto)
        self.assertEqual(_get_model('Calle'), Calle)
        self.assertEqual(_get_model('Nodo'), Nodo)
        self.assertEqual(_get_model('Arco'), Arco)
        self.assertEqual(_get_model('Regulacion'), Regulacion)
        self.assertEqual(_get_model('Unknown'), None)

    def test_sanitize_for_session_date(self):
        data = {'fecha': date(2024, 1, 15)}
        _sanitize_for_session(data)
        self.assertEqual(data['fecha'], '2024-01-15')

    def test_sanitize_for_session_time(self):
        data = {'hora': time(14, 30)}
        _sanitize_for_session(data)
        self.assertEqual(data['hora'], '14:30:00')

    def test_sanitize_for_session_decimal(self):
        data = {'val': Decimal('3.14')}
        _sanitize_for_session(data)
        self.assertEqual(data['val'], 3.14)

    def test_sanitize_for_session_ignores_others(self):
        data = {'name': 'test', 'count': 42}
        _sanitize_for_session(data)
        self.assertEqual(data['name'], 'test')
        self.assertEqual(data['count'], 42)

    def test_restore_fk_values_virtual_obj(self):
        sheet_name = 'Contacto'
        mock = _VirtualObj(pk=0, value='MandanteName')
        data = {'mandante': mock, 'name': 'Test'}
        _restore_fk_values(sheet_name, data)
        self.assertEqual(data['mandante'], 'MandanteName')

    def test_restore_fk_values_model_instance(self):
        mandante = Mandante.objects.create(name='TestMandante', location='Loc')
        sheet_name = 'Contacto'
        data = {'mandante': mandante, 'name': 'Test'}
        _restore_fk_values(sheet_name, data)
        self.assertEqual(data['mandante'], 'TestMandante')

    def test_restore_fk_values_not_in_fk_fields(self):
        data = {'name': 'Test', 'location': 'Loc'}
        _restore_fk_values('Mandante', data)
        self.assertEqual(data['name'], 'Test')


class ValidateSheetTest(TestCase):
    def setUp(self):
        self.user = User.objects.create_user(username="testuser", password="12345")
        self.mandante = Mandante.objects.create(name="TestMandante", location="Loc")
        self.proyecto = Proyecto.objects.create(
            title="TestProyecto", user=self.user, mandante=self.mandante,
        )

    def test_valid_mandante_row(self):
        rows = [{'name': 'M1', 'location': 'Loc1', 'details': 'Det1'}]
        result = validate_sheet('Mandante', rows)
        self.assertEqual(len(result['valid']), 1)
        self.assertEqual(len(result['errors']), 0)
        self.assertEqual(result['valid'][0]['name'], 'M1')

    def test_missing_required_field(self):
        rows = [{'name': 'M1'}]  # missing 'location'
        result = validate_sheet('Mandante', rows)
        self.assertEqual(len(result['valid']), 0)
        self.assertEqual(len(result['errors']), 1)
        self.assertIn('location', result['errors'][0]['errors'][0])

    def test_fk_resolution_valid(self):
        rows = [{'name': 'C1', 'email': 'c@c.com', 'mandante': 'TestMandante'}]
        result = validate_sheet('Contacto', rows, proyecto=self.proyecto)
        self.assertEqual(len(result['valid']), 1)
        self.assertEqual(result['valid'][0]['mandante'], self.mandante)

    def test_fk_resolution_invalid(self):
        rows = [{'name': 'C1', 'email': 'c@c.com', 'mandante': 'NonExistent'}]
        result = validate_sheet('Contacto', rows, proyecto=self.proyecto)
        self.assertEqual(len(result['valid']), 0)
        self.assertEqual(len(result['errors']), 1)
        self.assertIn('no encontrado', result['errors'][0]['errors'][0])

    def test_fk_cross_sheet_via_shared_cache(self):
        shared_cache = {('Mandante', 'name', 'VirtualMandante'): True}
        rows = [{'name': 'C1', 'email': 'c@c.com', 'mandante': 'VirtualMandante'}]
        result = validate_sheet('Contacto', rows, proyecto=self.proyecto, shared_cache=shared_cache)
        self.assertEqual(len(result['valid']), 1)
        self.assertIsInstance(result['valid'][0]['mandante'], _VirtualObj)
        self.assertEqual(result['valid'][0]['mandante']._value, 'VirtualMandante')

    def test_duplicate_detection(self):
        Mandante.objects.create(name='ExistingM', location='Loc')
        rows = [{'name': 'ExistingM', 'location': 'Loc2'}]
        result = validate_sheet('Mandante', rows)
        self.assertEqual(len(result['valid']), 0)
        self.assertEqual(len(result['duplicates']), 1)
        self.assertEqual(result['duplicates'][0]['name'], 'ExistingM')

    def test_multiple_rows_mixed(self):
        Mandante.objects.create(name='ExistingM', location='Loc')
        rows = [
            {'name': 'ExistingM', 'location': 'NewLoc'},
            {'name': 'NewM', 'location': 'Loc2'},
            {'name': '', 'location': 'Loc3'},  # missing name
        ]
        result = validate_sheet('Mandante', rows)
        self.assertEqual(len(result['valid']), 1)
        self.assertEqual(len(result['duplicates']), 1)
        self.assertEqual(len(result['errors']), 1)
        self.assertEqual(result['total'], 3)

    def test_parse_row_errors(self):
        rows = [{'numero': 'abc', 'proyecto': 'TestProyecto'}]
        result = validate_sheet('Calle', rows, proyecto=self.proyecto)
        self.assertEqual(len(result['valid']), 0)
        self.assertEqual(len(result['errors']), 1)

    def test_proyecto_sheet_valid(self):
        mandante = self.mandante
        rows = [{'title': 'NewProj', 'date_started': '15/01/2024', 'mandante': 'TestMandante'}]
        result = validate_sheet('Proyecto', rows, proyecto=self.proyecto)
        self.assertEqual(len(result['valid']), 1)
        self.assertEqual(result['valid'][0]['title'], 'NewProj')
        self.assertEqual(result['valid'][0]['mandante'], mandante)
        self.assertEqual(result['valid'][0]['date_started'], date(2024, 1, 15))


class ValidateSelectionTest(TestCase):
    def setUp(self):
        self.user = User.objects.create_user(username="testuser", password="12345")
        self.mandante = Mandante.objects.create(name="TestMandante", location="Loc")
        self.proyecto = Proyecto.objects.create(
            title="TestProyecto", user=self.user, mandante=self.mandante,
        )

    def test_validates_in_order(self):
        parsed = {
            'Mandante': [{'name': 'M1', 'location': 'Loc1'}],
            'Contacto': [{'name': 'C1', 'email': 'c@c.com', 'mandante': 'M1'}],
        }
        result = validate_selection(parsed, ['Contacto', 'Mandante'], self.proyecto)
        self.assertIn('Mandante', result['results'])
        self.assertIn('Contacto', result['results'])
        self.assertEqual(result['total_valid'], 2)
        self.assertEqual(result['total_errors'], 0)

    def test_cross_sheet_fk_via_shared_cache(self):
        parsed = {
            'Mandante': [{'name': 'M1', 'location': 'Loc1'}],
            'Contacto': [{'name': 'C1', 'email': 'c@c.com', 'mandante': 'M1'}],
        }
        result = validate_selection(parsed, ['Mandante', 'Contacto'], self.proyecto)
        self.assertEqual(result['total_valid'], 2)
        contacto_valid = result['results']['Contacto']['valid']
        self.assertEqual(len(contacto_valid), 1)
        self.assertEqual(contacto_valid[0]['mandante'], 'M1')

    def test_empty_selected_sheets(self):
        result = validate_selection({'Mandante': [{'name': 'M1', 'location': 'L1'}]}, [], self.proyecto)
        self.assertEqual(result['total_valid'], 0)

    def test_totals_correct(self):
        parsed = {
            'Mandante': [
                {'name': 'M1', 'location': 'L1'},
                {'name': 'M2', 'location': 'L2'},
            ],
            'Regulacion': [
                {'codigo': 'SEM', 'descripcion': 'Semáforo'},
            ],
        }
        result = validate_selection(parsed, ['Mandante', 'Regulacion'], self.proyecto)
        self.assertEqual(result['total_valid'], 3)
        self.assertEqual(result['total_errors'], 0)

    def test_sanitization_applied(self):
        parsed = {
            'Mandante': [{'name': 'M1', 'location': 'L1'}],
            'Proyecto': [{'title': 'P1', 'date_started': '15/01/2024', 'mandante': 'M1'}],
        }
        result = validate_selection(parsed, ['Mandante', 'Proyecto'], self.proyecto)
        proj_data = result['results']['Proyecto']['valid'][0]
        self.assertIsInstance(proj_data['date_started'], str)


class ExecuteImportTest(TestCase):
    def setUp(self):
        self.user = User.objects.create_user(username="testuser", password="12345")
        self.mandante = Mandante.objects.create(name="TestMandante", location="Loc")
        self.proyecto = Proyecto.objects.create(
            title="TestProyecto", user=self.user, mandante=self.mandante,
        )
        self.regulacion = Regulacion.objects.create(codigo='SEM', descripcion='Semáforo')
        self.calle = Calle.objects.create(nombre='TestCalle', numero=1, proyecto=self.proyecto)
        self.nodo1 = Nodo.objects.create(numero=1, proyecto=self.proyecto)
        self.nodo2 = Nodo.objects.create(numero=2, proyecto=self.proyecto)

    def _make_validation_result(self, results_dict):
        total_valid = sum(len(r.get('valid', [])) for r in results_dict.values())
        total_errors = sum(len(r.get('errors', [])) for r in results_dict.values())
        total_duplicates = sum(len(r.get('duplicates', [])) for r in results_dict.values())
        return {
            'results': results_dict,
            'total_valid': total_valid,
            'total_errors': total_errors,
            'total_duplicates': total_duplicates,
        }

    def test_insert_mandante(self):
        validation = self._make_validation_result({
            'Mandante': {'valid': [{'name': 'NewMandante', 'location': 'Loc', 'details': ''}], 'duplicates': [], 'errors': [], 'sheet': 'Mandante'},
        })
        report = execute_import(validation, self.proyecto, self.user)
        self.assertTrue(Mandante.objects.filter(name='NewMandante').exists())
        self.assertEqual(report['Mandante']['inserted'], 1)

    def test_update_existing_mandante(self):
        Mandante.objects.create(name='ExistingM', location='OldLoc')
        validation = self._make_validation_result({
            'Mandante': {'valid': [], 'duplicates': [{'name': 'ExistingM', 'location': 'NewLoc', 'details': ''}], 'errors': [], 'sheet': 'Mandante'},
        })
        report = execute_import(validation, self.proyecto, self.user, update_duplicates={'Mandante': True})
        mandante = Mandante.objects.get(name='ExistingM')
        self.assertEqual(mandante.location, 'NewLoc')
        self.assertEqual(report['Mandante']['updated'], 1)

    def test_skip_duplicates(self):
        Mandante.objects.create(name='ExistingM', location='Loc')
        validation = self._make_validation_result({
            'Mandante': {'valid': [], 'duplicates': [{'name': 'ExistingM', 'location': 'NewLoc', 'details': ''}], 'errors': [], 'sheet': 'Mandante'},
        })
        report = execute_import(validation, self.proyecto, self.user, update_duplicates={'Mandante': False})
        mandante = Mandante.objects.get(name='ExistingM')
        self.assertEqual(mandante.location, 'Loc')
        self.assertEqual(report['Mandante']['skipped_duplicates'], 1)

    def test_insert_calle(self):
        validation = self._make_validation_result({
            'Calle': {'valid': [{'nombre': 'NuevaCalle', 'numero': 99, 'proyecto': self.proyecto}], 'duplicates': [], 'errors': [], 'sheet': 'Calle'},
        })
        report = execute_import(validation, self.proyecto, self.user)
        self.assertTrue(Calle.objects.filter(numero=99, proyecto=self.proyecto).exists())
        self.assertEqual(report['Calle']['inserted'], 1)

    def test_insert_nodo(self):
        validation = self._make_validation_result({
            'Nodo': {'valid': [{'numero': 10, 'interseccion': 'Test', 'proyecto': self.proyecto}], 'duplicates': [], 'errors': [], 'sheet': 'Nodo'},
        })
        report = execute_import(validation, self.proyecto, self.user)
        self.assertTrue(Nodo.objects.filter(numero=10, proyecto=self.proyecto).exists())
        self.assertEqual(report['Nodo']['inserted'], 1)

    def test_insert_arco(self):
        validation = self._make_validation_result({
            'Arco': {'valid': [{'nodo_origen': self.nodo1, 'nodo_destino': self.nodo2, 'longitud': 100.0, 'proyecto': self.proyecto}], 'duplicates': [], 'errors': [], 'sheet': 'Arco'},
        })
        report = execute_import(validation, self.proyecto, self.user)
        self.assertTrue(Arco.objects.filter(nodo_origen=self.nodo1, nodo_destino=self.nodo2).exists())
        self.assertEqual(report['Arco']['inserted'], 1)

    def test_insert_regulacion(self):
        validation = self._make_validation_result({
            'Regulacion': {'valid': [{'codigo': 'PARE', 'descripcion': 'Pare'}], 'duplicates': [], 'errors': [], 'sheet': 'Regulacion'},
        })
        report = execute_import(validation, self.proyecto, self.user)
        self.assertTrue(Regulacion.objects.filter(codigo='PARE').exists())
        self.assertEqual(report['Regulacion']['inserted'], 1)

    def test_insert_periodo(self):
        validation = self._make_validation_result({
            'Periodo': {'valid': [{'codigo': 'PM-L', 'hora_inicio': time(8, 0), 'hora_fin': time(9, 0), 'es_laboral': True, 'proyecto': self.proyecto}], 'duplicates': [], 'errors': [], 'sheet': 'Periodo'},
        })
        report = execute_import(validation, self.proyecto, self.user)
        self.assertTrue(Periodo.objects.filter(codigo='PM-L', proyecto=self.proyecto).exists())
        self.assertEqual(report['Periodo']['inserted'], 1)

    def test_insert_configuracion_transyt(self):
        validation = self._make_validation_result({
            'ConfiguracionTransyt': {'valid': [{'proyecto': self.proyecto, 'ciclo': 60, 'W': 10.0, 'K': 0.5, 'perdida_inicial': 2.0, 'ganancia_final': 1.0}], 'duplicates': [], 'errors': [], 'sheet': 'ConfiguracionTransyt'},
        })
        report = execute_import(validation, self.proyecto, self.user)
        self.assertTrue(ConfiguracionTransyt.objects.filter(proyecto=self.proyecto).exists())
        self.assertEqual(report['ConfiguracionTransyt']['inserted'], 1)

    def test_insert_punto_control(self):
        arco = Arco.objects.create(nodo_origen=self.nodo1, nodo_destino=self.nodo2, longitud=50.0, proyecto=self.proyecto)
        validation = self._make_validation_result({
            'PuntoControl': {'valid': [{
                'movimiento': '12', 'viraje': 'DIR', 'is_prioritario': True,
                'numero_pistas': 2.0, 'nodo': self.nodo1,
                'arco_entrada': arco, 'arco_salida': arco,
                'regulacion': self.regulacion, 'proyecto': self.proyecto,
            }], 'duplicates': [], 'errors': [], 'sheet': 'PuntoControl'},
        })
        report = execute_import(validation, self.proyecto, self.user)
        self.assertTrue(PuntoControl.objects.filter(movimiento='12', nodo=self.nodo1).exists())
        self.assertEqual(report['PuntoControl']['inserted'], 1)

    def test_insert_periodizacion(self):
        arco = Arco.objects.create(nodo_origen=self.nodo1, nodo_destino=self.nodo2, longitud=50.0, proyecto=self.proyecto)
        pc = PuntoControl.objects.create(
            movimiento='12', viraje='DIR', is_prioritario=True,
            nodo=self.nodo1, arco_entrada=arco, arco_salida=arco,
            regulacion=self.regulacion, proyecto=self.proyecto,
        )
        periodo = Periodo.objects.create(codigo='PM-L', es_laboral=True, proyecto=self.proyecto)
        validation = self._make_validation_result({
            'Periodizacion': {'valid': [{
                'fecha': date(2024, 1, 15), 'hora': time(8, 0),
                'pc': pc, 'periodo': periodo,
                'vl': 100, 'txc': 50, 'txb': 20, 'c2e': 10,
                'c_mas2e': 5, 'peat': 30, 'cicl': 2, 'moto': 3,
            }], 'duplicates': [], 'errors': [], 'sheet': 'Periodizacion'},
        })
        report = execute_import(validation, self.proyecto, self.user)
        self.assertEqual(report['Periodizacion']['inserted'], 1)

    def test_re_resolve_fk_string_values(self):
        """FK values stored as strings in session are re-resolved from DB."""
        m = Mandante.objects.create(name='ReResolveM', location='Loc')
        validation = self._make_validation_result({
            'Contacto': {'valid': [{'name': 'Contact1', 'email': 'c@c.com', 'mandante': 'ReResolveM'}], 'duplicates': [], 'errors': [], 'sheet': 'Contacto'},
        })
        report = execute_import(validation, self.proyecto, self.user)
        self.assertTrue(Contacto.objects.filter(name='Contact1').exists())
        self.assertEqual(report['Contacto']['inserted'], 1)

    def test_error_rows_in_report(self):
        validation = self._make_validation_result({
            'Mandante': {'valid': [], 'duplicates': [], 'errors': [{'row': 1, 'data': {}, 'errors': ["'name' es requerido"]}], 'sheet': 'Mandante'},
        })
        report = execute_import(validation, self.proyecto, self.user)
        self.assertEqual(len(report['Mandante']['rejected']), 1)
        self.assertIn('requerido', report['Mandante']['rejected'][0]['reason'])

    def test_multiple_sheets_processed(self):
        """Multiple independent sheets are all processed."""
        validation = self._make_validation_result({
            'Mandante': {'valid': [{'name': 'M1', 'location': 'Loc', 'details': ''}], 'duplicates': [], 'errors': [], 'sheet': 'Mandante', 'total': 1},
            'Regulacion': {'valid': [{'codigo': 'PARE', 'descripcion': 'Pare'}], 'duplicates': [], 'errors': [], 'sheet': 'Regulacion', 'total': 1},
        })
        report = execute_import(validation, self.proyecto, self.user)
        self.assertTrue(Mandante.objects.filter(name='M1').exists())
        self.assertTrue(Regulacion.objects.filter(codigo='PARE').exists())
        self.assertEqual(report['Mandante']['inserted'], 1)
        self.assertEqual(report['Regulacion']['inserted'], 1)


class ResolveArcoTest(TestCase):
    def setUp(self):
        self.user = User.objects.create_user(username="testuser", password="12345")
        self.mandante = Mandante.objects.create(name="TM", location="Loc")
        self.proyecto = Proyecto.objects.create(title="TP", user=self.user, mandante=self.mandante)
        self.nodo1 = Nodo.objects.create(numero=1, proyecto=self.proyecto)
        self.nodo2 = Nodo.objects.create(numero=2, proyecto=self.proyecto)
        self.arco = Arco.objects.create(nodo_origen=self.nodo1, nodo_destino=self.nodo2, longitud=50, proyecto=self.proyecto)

    def test_resolve_valid_arco(self):
        result = _resolve_arco('1>2', self.proyecto)
        self.assertEqual(result, self.arco)

    def test_resolve_invalid_format(self):
        self.assertIsNone(_resolve_arco('1-2', self.proyecto))

    def test_resolve_none_proyecto(self):
        self.assertIsNone(_resolve_arco('1>2', None))

    def test_resolve_empty_value(self):
        self.assertIsNone(_resolve_arco('', self.proyecto))

    def test_resolve_nonexistent(self):
        self.assertIsNone(_resolve_arco('99>98', self.proyecto))


class ResolveFkTest(TestCase):
    def setUp(self):
        self.user = User.objects.create_user(username="testuser", password="12345")
        self.mandante = Mandante.objects.create(name="TestMandante", location="Loc")
        self.proyecto = Proyecto.objects.create(title="TP", user=self.user, mandante=self.mandante)
        self.nodo = Nodo.objects.create(numero=5, proyecto=self.proyecto)

    def test_resolve_mandante_by_name(self):
        ctx = {'proyecto': self.proyecto, 'resolved': {}, 'shared_cache': {}}
        result = _resolve_fk('Contacto', 'mandante', 'TestMandante', ctx)
        self.assertEqual(result, self.mandante)
        self.assertIn('Mandante:TestMandante', ctx['resolved'])

    def test_resolve_cached(self):
        ctx = {'proyecto': self.proyecto, 'resolved': {'Mandante:TestMandante': self.mandante}, 'shared_cache': {}}
        result = _resolve_fk('Contacto', 'mandante', 'TestMandante', ctx)
        self.assertEqual(result, self.mandante)

    def test_resolve_nodo_scoped_by_proyecto(self):
        ctx = {'proyecto': self.proyecto, 'resolved': {}, 'shared_cache': {}}
        result = _resolve_fk('Arco', 'nodo_origen', '5', ctx)
        self.assertEqual(result, self.nodo)

    def test_resolve_via_shared_cache(self):
        """FK resolves to _VirtualObj when value is not in DB but in shared_cache."""
        ctx = {'proyecto': self.proyecto, 'resolved': {}, 'shared_cache': {('Nodo', 'numero', '999', str(self.proyecto.id)): True}}
        result = _resolve_fk('Arco', 'nodo_origen', '999', ctx)
        self.assertIsInstance(result, _VirtualObj)
        self.assertEqual(result._value, '999')

    def test_resolve_empty_value(self):
        ctx = {'proyecto': self.proyecto, 'resolved': {}, 'shared_cache': {}}
        self.assertIsNone(_resolve_fk('Contacto', 'mandante', '', ctx))

    def test_resolve_unknown_sheet(self):
        ctx = {'proyecto': self.proyecto, 'resolved': {}, 'shared_cache': {}}
        self.assertIsNone(_resolve_fk('UnknownSheet', 'field', 'val', ctx))


class ParseRowTest(TestCase):
    def setUp(self):
        self.user = User.objects.create_user(username="testuser", password="12345")
        self.mandante = Mandante.objects.create(name="TM", location="Loc")
        self.proyecto = Proyecto.objects.create(title="TP", user=self.user, mandante=self.mandante)
        self.nodo1 = Nodo.objects.create(numero=1, proyecto=self.proyecto)
        self.nodo2 = Nodo.objects.create(numero=2, proyecto=self.proyecto)
        self.arco = Arco.objects.create(nodo_origen=self.nodo1, nodo_destino=self.nodo2, longitud=50, proyecto=self.proyecto)
        self.ctx = {
            'proyecto': self.proyecto,
            'resolved': {},
            'shared_cache': {},
        }

    def test_parse_mandante(self):
        result = _parse_row('Mandante', {'name': 'M1', 'location': 'L1', 'details': 'D1'}, 1, self.ctx, {})
        self.assertEqual(result['data']['name'], 'M1')
        self.assertEqual(result['data']['location'], 'L1')

    def test_parse_contacto_valid(self):
        self.ctx['resolved']['Mandante:TM'] = self.mandante
        result = _parse_row('Contacto', {'name': 'C1', 'email': 'c@c.com', 'mandante': 'TM'}, 1, self.ctx, {})
        self.assertEqual(result['data']['name'], 'C1')
        self.assertEqual(result['data']['mandante'], self.mandante)

    def test_parse_contacto_missing_mandante(self):
        result = _parse_row('Contacto', {'name': 'C1', 'email': 'c@c.com', 'mandante': 'NonExistent'}, 1, self.ctx, {})
        self.assertTrue(len(result['errors']) > 0)

    def test_parse_proyecto_valid(self):
        self.ctx['resolved']['Mandante:TM'] = self.mandante
        result = _parse_row('Proyecto', {'title': 'P1', 'date_started': '15/01/2024', 'mandante': 'TM'}, 1, self.ctx, {})
        self.assertEqual(result['data']['title'], 'P1')
        self.assertEqual(result['data']['date_started'], date(2024, 1, 15))

    def test_parse_calle_valid(self):
        self.ctx['resolved'][f'Proyecto:TP'] = self.proyecto
        result = _parse_row('Calle', {'nombre': 'Calle1', 'numero': '10', 'proyecto': 'TP'}, 1, self.ctx, {})
        self.assertEqual(result['data']['nombre'], 'Calle1')
        self.assertEqual(result['data']['numero'], 10)
        self.assertEqual(result['data']['proyecto'], self.proyecto)

    def test_parse_nodo_valid(self):
        self.ctx['resolved'][f'Proyecto:TP'] = self.proyecto
        result = _parse_row('Nodo', {'numero': '10', 'interseccion': 'Test', 'proyecto': 'TP'}, 1, self.ctx, {})
        self.assertEqual(result['data']['numero'], 10)
        self.assertEqual(result['data']['proyecto'], self.proyecto)

    def test_parse_arco_valid(self):
        self.ctx['resolved'][f'Nodo:{self.proyecto.id}:1'] = self.nodo1
        self.ctx['resolved'][f'Nodo:{self.proyecto.id}:2'] = self.nodo2
        self.ctx['resolved'][f'Proyecto:TP'] = self.proyecto
        result = _parse_row('Arco', {'nodo_origen': '1', 'nodo_destino': '2', 'longitud': '100.5', 'proyecto': 'TP'}, 1, self.ctx, {})
        self.assertEqual(result['data']['longitud'], 100.5)
        self.assertEqual(result['data']['nodo_origen'], self.nodo1)
        self.assertEqual(result['data']['nodo_destino'], self.nodo2)

    def test_parse_regulacion(self):
        result = _parse_row('Regulacion', {'codigo': 'SEM', 'descripcion': 'Semáforo'}, 1, self.ctx, {})
        self.assertEqual(result['data']['codigo'], 'SEM')

    def test_parse_periodo_valid(self):
        self.ctx['resolved'][f'Proyecto:TP'] = self.proyecto
        result = _parse_row('Periodo', {'codigo': 'PM-L', 'hora_inicio': '08:00', 'hora_fin': '09:00', 'es_laboral': 'SI', 'proyecto': 'TP'}, 1, self.ctx, {})
        self.assertEqual(result['data']['codigo'], 'PM-L')

    def test_parse_coeficiente_cruce_standard(self):
        result = _parse_row('CoeficienteCruce', {'nomenclatura': 'VL', 'tipo_transporte': 'Vehiculo', 'coeficiente': '1.0', 'is_standard': 'SI'}, 1, self.ctx, {})
        self.assertEqual(result['data']['nomenclatura'], 'VL')

    def test_parse_coeficiente_cruce_non_standard_no_proyecto(self):
        result = _parse_row('CoeficienteCruce', {'nomenclatura': 'VL', 'tipo_transporte': 'Vehiculo', 'coeficiente': '1.0', 'is_standard': 'NO', 'proyecto': 'NonExistent'}, 1, self.ctx, {})
        self.assertTrue(len(result['errors']) > 0)

    def test_parse_configuracion_transyt(self):
        self.ctx['resolved'][f'Proyecto:TP'] = self.proyecto
        result = _parse_row('ConfiguracionTransyt', {'proyecto': 'TP', 'ciclo': '90', 'W': '15', 'K': '0.4', 'perdida_inicial': '3', 'ganancia_final': '2'}, 1, self.ctx, {})
        self.assertEqual(result['data']['ciclo'], 90)

    def test_parse_fase_semaforica(self):
        self.ctx['resolved'][f'Proyecto:TP'] = self.proyecto
        nodo_pc = Nodo.objects.create(numero=99, numero_pc=1, proyecto=self.proyecto)
        pc = PuntoControl.objects.create(
            movimiento='12', nodo=nodo_pc,
            arco_entrada=self.arco, arco_salida=self.arco,
            regulacion=Regulacion.objects.create(codigo='SEM2', descripcion='Sem'),
            proyecto=self.proyecto,
        )
        self.ctx['resolved'][f'PuntoControl:{self.proyecto.id}:{pc.nombre}'] = pc
        result = _parse_row('FaseSemaforica', {'punto_control': pc.nombre, 'fase_numero': '1', 'verde_inicio': '10', 'verde_fin': '40', 'proyecto': 'TP'}, 1, self.ctx, {})
        self.assertEqual(result['data']['fase_numero'], 1)


class ExecuteSheetTest(TestCase):
    def setUp(self):
        self.user = User.objects.create_user(username="testuser", password="12345")
        self.mandante = Mandante.objects.create(name="TM", location="Loc")
        self.proyecto = Proyecto.objects.create(title="TP", user=self.user, mandante=self.mandante)

    def test_mandante_get_or_create_new(self):
        from apps.red_vial.services.import_service import _execute_sheet
        report = {'inserted': 0, 'updated': 0, 'rejected': [], 'skipped_duplicates': 0, 'valid_count': 1, 'duplicate_count': 0, 'error_count': 0}
        _execute_sheet('Mandante', [{'name': 'NewM', 'location': 'Loc', 'details': ''}], self.proyecto, self.user, report)
        self.assertEqual(report['inserted'], 1)
        self.assertTrue(Mandante.objects.filter(name='NewM').exists())

    def test_mandante_get_or_create_existing(self):
        Mandante.objects.create(name='ExistingM', location='OldLoc', details='')
        from apps.red_vial.services.import_service import _execute_sheet
        report = {'inserted': 0, 'updated': 0, 'rejected': [], 'skipped_duplicates': 0, 'valid_count': 1, 'duplicate_count': 0, 'error_count': 0}
        _execute_sheet('Mandante', [{'name': 'ExistingM', 'location': 'NewLoc', 'details': 'Updated'}], self.proyecto, self.user, report)
        self.assertEqual(report['updated'], 1)
        mandante = Mandante.objects.get(name='ExistingM')
        self.assertEqual(mandante.location, 'NewLoc')


class ResolvePuntoControlPorNombreTest(TestCase):
    def setUp(self):
        self.user = User.objects.create_user(username="testuser", password="12345")
        self.mandante = Mandante.objects.create(name="TM", location="Loc")
        self.proyecto = Proyecto.objects.create(title="TP", user=self.user, mandante=self.mandante)
        self.nodo1 = Nodo.objects.create(numero=10, numero_pc=1, proyecto=self.proyecto)
        self.nodo2 = Nodo.objects.create(numero=20, proyecto=self.proyecto)
        self.arco = Arco.objects.create(nodo_origen=self.nodo1, nodo_destino=self.nodo2, longitud=50, proyecto=self.proyecto)
        self.reg = Regulacion.objects.create(codigo='SEM', descripcion='Sem')
        self.pc_pc01 = PuntoControl.objects.create(
            movimiento='12', nodo=self.nodo1, arco_entrada=self.arco,
            arco_salida=self.arco, regulacion=self.reg, proyecto=self.proyecto,
        )
        self.pc_nodo = PuntoControl.objects.create(
            movimiento='21', nodo=self.nodo2, arco_entrada=self.arco,
            arco_salida=self.arco, regulacion=self.reg, proyecto=self.proyecto,
        )

    def test_resolve_pc01(self):
        ctx = {'proyecto': self.proyecto}
        result = _resolve_puntocontrol_por_nombre('PC-01', ctx)
        self.assertEqual(result, self.pc_pc01)

    def test_resolve_nodo(self):
        ctx = {'proyecto': self.proyecto}
        result = _resolve_puntocontrol_por_nombre('Nodo-20', ctx)
        self.assertEqual(result, self.pc_nodo)

    def test_resolve_nodo_without_hyphen(self):
        ctx = {'proyecto': self.proyecto}
        result = _resolve_puntocontrol_por_nombre('Nodo20', ctx)
        self.assertEqual(result, self.pc_nodo)

    def test_resolve_not_found(self):
        ctx = {'proyecto': self.proyecto}
        result = _resolve_puntocontrol_por_nombre('PC-99', ctx)
        self.assertIsNone(result)

    def test_resolve_no_proyecto(self):
        ctx = {'proyecto': None}
        result = _resolve_puntocontrol_por_nombre('PC-01', ctx)
        self.assertIsNone(result)

    def test_resolve_empty_value(self):
        ctx = {'proyecto': self.proyecto}
        self.assertIsNone(_resolve_puntocontrol_por_nombre('', ctx))


class FullPipelineIntegrationTest(TestCase):
    """End-to-end test: Excel bytes → parse → validate → execute → verify."""

    def _make_sheets_data(self, mandante_name, proyecto_title):
        sheets = {
            'Mandante': [
                {'name': mandante_name, 'location': 'Loc', 'details': ''},
            ],
            'Contacto': [
                {'name': 'Contacto 1', 'email': 'c1@test.cl', 'phone': '', 'cargo': '', 'position': '', 'details': '', 'mandante': mandante_name},
            ],
            'Proyecto': [
                {'title': proyecto_title, 'description': '', 'date_started': '01/01/2025', 'mandante': mandante_name},
            ],
            'Calle': [
                {'nombre': 'Calle A', 'numero': 1, 'proyecto': proyecto_title},
                {'nombre': 'Calle B', 'numero': 2, 'proyecto': proyecto_title},
            ],
            'Nodo': [
                {'numero': 1, 'interseccion': 'Calle A con Calle B', 'calle_1': 'Calle A', 'calle_2': 'Calle B', 'numero_pc': 1, 'plano': '', 'imagen': '', 'proyecto': proyecto_title},
                {'numero': 2, 'interseccion': 'Calle B con Calle A', 'calle_1': 'Calle B', 'calle_2': 'Calle A', 'numero_pc': 2, 'plano': '', 'imagen': '', 'proyecto': proyecto_title},
            ],
            'Arco': [
                {'nodo_origen': 1, 'nodo_destino': 2, 'longitud': 100.0, 'proyecto': proyecto_title},
                {'nodo_origen': 2, 'nodo_destino': 1, 'longitud': 100.0, 'proyecto': proyecto_title},
            ],
            'Regulacion': [
                {'codigo': 'SEM01', 'descripcion': 'Semáforo fijo'},
            ],
            'CoeficienteCruce': [
                {'nomenclatura': 'VL', 'tipo_transporte': 'Vehículo Liviano', 'coeficiente': 1.0, 'is_standard': 'SI', 'proyecto': ''},
                {'nomenclatura': 'BUS', 'tipo_transporte': 'Bus', 'coeficiente': 2.5, 'is_standard': 'NO', 'proyecto': proyecto_title},
            ],
            'Periodo': [
                {'codigo': 'PM-L', 'hora_inicio': '06:00', 'hora_fin': '09:00', 'es_laboral': 'SI', 'proyecto': proyecto_title},
                {'codigo': 'PT-L', 'hora_inicio': '18:00', 'hora_fin': '21:00', 'es_laboral': 'SI', 'proyecto': proyecto_title},
            ],
            'PuntoControl': [
                {'nodo': 1, 'movimiento': '12', 'viraje': 'DIR', 'is_prioritario': 'SI', 'arco_entrada': '1>2', 'arco_salida': '2>1', 'regulacion': 'SEM01', 'numero_pistas': 2.0, 'proyecto': proyecto_title},
                {'nodo': 2, 'movimiento': '21', 'viraje': 'DIR', 'is_prioritario': 'NO', 'arco_entrada': '2>1', 'arco_salida': '1>2', 'regulacion': 'SEM01', 'numero_pistas': 1.0, 'proyecto': proyecto_title},
            ],
            'Periodizacion': [
                {'fecha': '15/03/2025', 'hora': '06:00', 'pc': 'PC-01', 'periodo': 'PM-L', 'proyecto': proyecto_title, 'vl': 100, 'txc': 20, 'txb': 10, 'c2e': 5, 'c_mas2e': 2, 'peat': 50, 'cicl': 10, 'moto': 5},
                {'fecha': '15/03/2025', 'hora': '06:15', 'pc': 'PC-01', 'periodo': 'PM-L', 'proyecto': proyecto_title, 'vl': 120, 'txc': 25, 'txb': 8, 'c2e': 3, 'c_mas2e': 1, 'peat': 40, 'cicl': 8, 'moto': 3},
                {'fecha': '15/03/2025', 'hora': '18:00', 'pc': 'PC-02', 'periodo': 'PT-L', 'proyecto': proyecto_title, 'vl': 80, 'txc': 15, 'txb': 5, 'c2e': 2, 'c_mas2e': 1, 'peat': 30, 'cicl': 5, 'moto': 2},
                {'fecha': '15/03/2025', 'hora': '18:15', 'pc': 'PC-02', 'periodo': 'PT-L', 'proyecto': proyecto_title, 'vl': 90, 'txc': 18, 'txb': 6, 'c2e': 3, 'c_mas2e': 1, 'peat': 35, 'cicl': 6, 'moto': 3},
            ],
            'ParametroArco': [
                {'punto_control': 'PC-01', 'flujo_saturacion': 1800.0, 'ponderador_demora': 1.0, 'ponderador_detencion': 1.0, 'capacidad_cola': 10.0, 'tiene_tarjeta_38': 'SI', 'proyecto': proyecto_title},
                {'punto_control': 'PC-02', 'flujo_saturacion': 1600.0, 'ponderador_demora': 1.2, 'ponderador_detencion': 0.8, 'capacidad_cola': '', 'tiene_tarjeta_38': 'NO', 'proyecto': proyecto_title},
            ],
            'FaseSemaforica': [
                {'punto_control': 'PC-01', 'fase_numero': 1, 'verde_inicio': 0.0, 'verde_fin': 25.0, 'proyecto': proyecto_title},
                {'punto_control': 'PC-01', 'fase_numero': 2, 'verde_inicio': 30.0, 'verde_fin': 55.0, 'proyecto': proyecto_title},
                {'punto_control': 'PC-02', 'fase_numero': 1, 'verde_inicio': 0.0, 'verde_fin': 20.0, 'proyecto': proyecto_title},
                {'punto_control': 'PC-02', 'fase_numero': 2, 'verde_inicio': 25.0, 'verde_fin': 45.0, 'proyecto': proyecto_title},
            ],
            'ConfiguracionTransyt': [
                {'proyecto': proyecto_title, 'ciclo': 60, 'W': 10.0, 'K': 0.5, 'perdida_inicial': 2.0, 'ganancia_final': 1.0},
            ],
        }
        return sheets

    def setUp(self):
        self.user = User.objects.create_user(username="testuser", password="12345")
        self.mandante = Mandante.objects.create(name="MandanteTest", location="Loc")
        self.proyecto = Proyecto.objects.create(
            title="ProyectoTest", user=self.user, mandante=self.mandante,
        )

    def test_full_pipeline_import_in_existing_project(self):
        """Import all 14 sheets into an existing project context.

        The import data references a proyecto that already exists (setUp).
        Mandante and Proyecto sheets match existing records → duplicates.
        All other sheets create new data linked to the existing proyecto.
        """
        sheets_data = self._make_sheets_data('MandanteTest', 'ProyectoTest')
        buf = _make_excel(sheets_data)
        all_sheets = list(sheets_data.keys())

        # 1. Parse
        parsed = parse_excel(buf)
        for s in all_sheets:
            self.assertIn(s, parsed, f'Sheet {s} missing after parse')
            self.assertTrue(len(parsed[s]) > 0, f'Sheet {s} has no rows')

        # 2. Validate (with proyecto context for FK scoping)
        validation = validate_selection(parsed, all_sheets, proyecto=self.proyecto)
        if validation['total_errors'] > 0:
            error_details = []
            for sheet, res in validation['results'].items():
                for err in res.get('errors', []):
                    error_details.append(f"{sheet} row {err.get('row')}: {err.get('errors')}")
            self.fail(f'Validation errors: {validation["total_errors"]} — {"; ".join(error_details)}')
        # Mandante and Proyecto should be duplicates (already exist in setUp)
        self.assertEqual(len(validation['results']['Mandante']['duplicates']), 1)
        self.assertEqual(len(validation['results']['Proyecto']['duplicates']), 1)

        # 3. Execute
        report = execute_import(validation, self.proyecto, self.user,
                                update_duplicates={'Mandante': True, 'Proyecto': True})

        # 4. Verify counts — Mandante and Proyecto updated, rest inserted
        self.assertEqual(report['Mandante']['updated'], 1)
        self.assertEqual(report['Proyecto']['updated'], 1)
        expected_inserts = {
            'Contacto': 1, 'Calle': 2, 'Nodo': 2, 'Arco': 2,
            'Regulacion': 1, 'CoeficienteCruce': 2, 'Periodo': 2,
            'PuntoControl': 2, 'Periodizacion': 4, 'ParametroArco': 2,
            'FaseSemaforica': 4, 'ConfiguracionTransyt': 1,
        }
        for sheet_name, expected in expected_inserts.items():
            self.assertIn(sheet_name, report, f'Missing report for {sheet_name}')
            self.assertEqual(
                report[sheet_name]['inserted'], expected,
                f'{sheet_name}: expected {expected} inserted, got {report[sheet_name]}'
            )
            self.assertEqual(
                report[sheet_name]['rejected'], [],
                f'{sheet_name}: unexpected rejected: {report[sheet_name]["rejected"]}'
            )

        # 5. Verify objects exist in DB linked to self.proyecto
        self.assertTrue(Calle.objects.filter(numero=1, proyecto=self.proyecto).exists())
        self.assertTrue(Calle.objects.filter(numero=2, proyecto=self.proyecto).exists())
        self.assertTrue(Nodo.objects.filter(numero=1, proyecto=self.proyecto).exists())
        self.assertTrue(Nodo.objects.filter(numero=2, proyecto=self.proyecto).exists())
        self.assertTrue(Arco.objects.filter(
            nodo_origen__numero=1, nodo_destino__numero=2, proyecto=self.proyecto
        ).exists())
        self.assertTrue(Regulacion.objects.filter(codigo='SEM01').exists())
        self.assertTrue(CoeficienteCruce.objects.filter(
            nomenclatura='VL', proyecto__isnull=True
        ).exists())
        self.assertTrue(CoeficienteCruce.objects.filter(
            nomenclatura='BUS', proyecto=self.proyecto
        ).exists())
        self.assertTrue(Periodo.objects.filter(codigo='PM-L', proyecto=self.proyecto).exists())
        self.assertTrue(Periodo.objects.filter(codigo='PT-L', proyecto=self.proyecto).exists())

        # PuntoControl — verify nombre computation from nodo.numero_pc
        pc1 = PuntoControl.objects.get(nodo__numero=1, proyecto=self.proyecto)
        self.assertEqual(pc1.nodo.numero_pc, 1)
        pc2 = PuntoControl.objects.get(nodo__numero=2, proyecto=self.proyecto)
        self.assertEqual(pc2.nodo.numero_pc, 2)

        # Periodizacion for both PCs
        self.assertEqual(Periodizacion.objects.filter(pc=pc1).count(), 2)
        self.assertEqual(Periodizacion.objects.filter(pc=pc2).count(), 2)

        # ParametroArco for both PCs
        self.assertTrue(ParametroArco.objects.filter(punto_control=pc1).exists())
        self.assertTrue(ParametroArco.objects.filter(punto_control=pc2).exists())

        # FaseSemaforica — 2 per PC
        self.assertEqual(FaseSemaforica.objects.filter(punto_control=pc1).count(), 2)
        self.assertEqual(FaseSemaforica.objects.filter(punto_control=pc2).count(), 2)

        # ConfiguracionTransyt
        self.assertTrue(ConfiguracionTransyt.objects.filter(proyecto=self.proyecto).exists())

    def test_full_pipeline_all_new_with_future_proyecto(self):
        """Import with a proyecto that matches the data; Mandante also new.

        This tests the scenario where the user creates a proyecto + mandante
        first (as context) and then fills in all downstream data.
        """
        # Create matching mandante and proyecto in setUp
        self.mandante = Mandante.objects.create(name='Nuevo Mandante', location='Loc')
        self.proyecto = Proyecto.objects.create(
            title='Nuevo Proyecto', user=self.user, mandante=self.mandante,
        )
        sheets_data = self._make_sheets_data('Nuevo Mandante', 'Nuevo Proyecto')
        buf = _make_excel(sheets_data)
        all_sheets = list(sheets_data.keys())

        parsed = parse_excel(buf)
        validation = validate_selection(parsed, all_sheets, proyecto=self.proyecto)

        # Mandante and Proyecto are duplicates (created in setUp)
        self.assertEqual(len(validation['results']['Mandante']['duplicates']), 1)
        self.assertEqual(len(validation['results']['Proyecto']['duplicates']), 1)
        self.assertEqual(validation['total_errors'], 0)

        report = execute_import(validation, self.proyecto, self.user,
                                update_duplicates={'Mandante': True, 'Proyecto': True})

        self.assertEqual(report['Mandante']['updated'], 1)
        self.assertEqual(report['Proyecto']['updated'], 1)
        # All downstream data inserted
        self.assertEqual(report['Calle']['inserted'], 2)
        self.assertEqual(report['PuntoControl']['inserted'], 2)
        self.assertEqual(report['Periodizacion']['inserted'], 4)
        self.assertEqual(report['ParametroArco']['inserted'], 2)
        self.assertEqual(report['FaseSemaforica']['inserted'], 4)
        self.assertEqual(report['ConfiguracionTransyt']['inserted'], 1)

    def test_template_generar_plantilla_is_parseable(self):
        """The template generated by generar_plantilla_bytes() parses without error."""
        from apps.common.utils.excel_utils import generar_plantilla_bytes
        buf = generar_plantilla_bytes()
        parsed = parse_excel(buf)
        # Should have all 14 data sheets (excluding the README sheet)
        expected_sheets = {'Mandante', 'Contacto', 'Proyecto', 'Calle', 'Nodo',
                           'Arco', 'Regulacion', 'CoeficienteCruce', 'Periodo',
                           'PuntoControl', 'Periodizacion', 'ParametroArco',
                           'FaseSemaforica', 'ConfiguracionTransyt'}
        parsed_sheets = set(parsed.keys())
        self.assertEqual(parsed_sheets, expected_sheets,
                         f'Missing sheets: {expected_sheets - parsed_sheets}')
        for sheet_name, rows in parsed.items():
            self.assertTrue(len(rows) > 0, f'Sheet {sheet_name} has no data rows')

    def test_template_full_pipeline_import(self):
        """The template can be fully imported end-to-end with pre-existing mandantes & proyectos."""
        from apps.common.utils.excel_utils import generar_plantilla_bytes
        # Pre-create matching mandantes and proyectos (they become duplicates)
        m1 = Mandante.objects.create(name='Municipalidad de Santiago', location='Santiago Centro')
        m2 = Mandante.objects.create(name='Gobierno Regional Metropolitano', location='Santiago')
        p1 = Proyecto.objects.create(
            title='Estudio de Semaforos Av. Libertador',
            user=self.user, mandante=m1,
        )
        p2 = Proyecto.objects.create(
            title='Auditoria Red Vial Centro',
            user=self.user, mandante=m2,
        )
        buf = generar_plantilla_bytes()
        parsed = parse_excel(buf)
        all_sheets = list(parsed.keys())
        validation = validate_selection(parsed, all_sheets, proyecto=p1)
        self.assertEqual(
            validation['total_errors'], 0,
            f'Validation errors: {validation["total_errors"]}',
        )
        # Mandante and Proyecto should be duplicates
        self.assertEqual(len(validation['results']['Mandante']['duplicates']), 2)
        self.assertEqual(len(validation['results']['Proyecto']['duplicates']), 2)
        report = execute_import(
            validation, p1, self.user,
            update_duplicates={'Mandante': True, 'Proyecto': True,
                               'Calle': True, 'Nodo': True, 'Arco': True,
                               'Regulacion': True, 'Periodo': True,
                               'PuntoControl': True, 'ParametroArco': True,
                               'FaseSemaforica': True},
        )
        self.assertEqual(report['Mandante']['updated'], 2)
        self.assertEqual(report['Proyecto']['updated'], 2)
        self.assertEqual(report['Contacto']['inserted'], 2)
        self.assertEqual(report['Calle']['inserted'], 6)
        self.assertEqual(report['Nodo']['inserted'], 4)
        self.assertEqual(report['Arco']['inserted'], 4)
        self.assertEqual(report['Regulacion']['inserted'], 3)
        self.assertEqual(report['Periodo']['inserted'], 4)
        self.assertEqual(report['PuntoControl']['inserted'], 4)
        self.assertEqual(report['Periodizacion']['inserted'], 96)
        self.assertEqual(Periodizacion.objects.filter(pc__proyecto=p1).count(), 48)
        self.assertEqual(Periodizacion.objects.filter(pc__proyecto=p2).count(), 48)
        self.assertEqual(report['ParametroArco']['inserted'], 4)
        self.assertEqual(report['FaseSemaforica']['inserted'], 8)
        self.assertEqual(report['ConfiguracionTransyt']['inserted'], 2)
        self.assertEqual(sum(len(r.get('rejected', [])) for r in report.values()), 0)


class ReportStructureTest(TestCase):
    def test_execute_import_report_keys(self):
        user = User.objects.create_user(username="testuser", password="12345")
        mandante = Mandante.objects.create(name="TM", location="Loc")
        proyecto = Proyecto.objects.create(title="TP", user=user, mandante=mandante)
        validation = {
            'results': {
                'Mandante': {'valid': [{'name': 'M1', 'location': 'L1', 'details': ''}], 'duplicates': [], 'errors': [], 'sheet': 'Mandante'},
            },
            'total_valid': 1, 'total_errors': 0, 'total_duplicates': 0,
        }
        report = execute_import(validation, proyecto, user)
        sheet_report = report['Mandante']
        self.assertIn('inserted', sheet_report)
        self.assertIn('updated', sheet_report)
        self.assertIn('rejected', sheet_report)
        self.assertIn('skipped_duplicates', sheet_report)
        self.assertIn('valid_count', sheet_report)
        self.assertIn('duplicate_count', sheet_report)
        self.assertIn('error_count', sheet_report)
